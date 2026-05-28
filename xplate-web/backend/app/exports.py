from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


EXPORT_DIR = Path(__file__).resolve().parents[1] / "exports"
EXPORT_DIR.mkdir(exist_ok=True)


def safe_name(value: str) -> str:
    return "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in value).strip("_") or "export"


def export_csv(rows: list[dict[str, Any]], filename_prefix: str = "xplate_results") -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    path = EXPORT_DIR / f"{safe_name(filename_prefix)}_{timestamp}.csv"
    pd.DataFrame(rows).to_csv(path, index=False, encoding="utf-8-sig")
    return str(path)


def export_excel(rows: list[dict[str, Any]], filename_prefix: str = "xplate_results") -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    path = EXPORT_DIR / f"{safe_name(filename_prefix)}_{timestamp}.xlsx"
    pd.DataFrame(rows).to_excel(path, index=False)
    workbook = load_workbook(path)
    sheet = workbook.active
    sheet.freeze_panes = "A2"
    fill = PatternFill(fill_type="solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
    for cells in sheet.columns:
        width = min(max(max(len(str(cell.value or "")) for cell in cells) + 2, 12), 55)
        sheet.column_dimensions[cells[0].column_letter].width = width
    workbook.save(path)
    return str(path)
