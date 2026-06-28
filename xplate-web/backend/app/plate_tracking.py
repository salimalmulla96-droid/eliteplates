"""
Plate Tracking and Duplicate Detection Module

Tracks unique plates across releases and manages cooldown periods to prevent
duplicate Telegram alerts for the same listing.
"""

import json
import hashlib
import re
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Optional
import logging
from zoneinfo import ZoneInfo

from .storage import DATA_DIR

logger = logging.getLogger(__name__)

# Path to the database
DB_PATH = DATA_DIR / 'plate_tracking.db'
REPORTS_DIR = Path(__file__).resolve().parents[1] / "reports" / "daily"
RULE_REPORTS_DIR = Path(__file__).resolve().parents[1] / "reports" / "daily_rules"
MISSING_PRICE_VALUES = {"", "?", "n/a", "na", "none", "null", "not available", "-"}


def daily_report_filename(date_str: str) -> str:
    return f"XPLATE REPORT {date_str}.xlsx"


def _ensure_columns(cursor: sqlite3.Cursor, table_name: str, columns: dict[str, str]) -> None:
    """Add any missing columns for lightweight SQLite migrations."""
    existing = {row[1] for row in cursor.execute(f"PRAGMA table_info({table_name})").fetchall()}
    for column_name, column_definition in columns.items():
        if column_name not in existing:
            cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_definition}")


def _is_missing_price(value: Any) -> bool:
    return str(value or "").strip().lower() in MISSING_PRICE_VALUES


def _price_for_report(value: Any) -> str:
    text = str(value or "").strip()
    return "N/A" if _is_missing_price(text) else text


def _parse_report_date(date_str: str) -> datetime:
    return datetime.strptime(str(date_str or "").strip(), "%Y-%m-%d")


def _safe_sheet_title(value: str, used: set[str] | None = None) -> str:
    used = used if used is not None else set()
    title = re.sub(r"[\[\]\:\*\?\/\\]", " ", str(value or "").strip())
    title = re.sub(r"\s+", " ", title).strip() or "Sheet"
    title = title[:31]
    base = title
    counter = 2
    while title in used:
        suffix = f" {counter}"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(title)
    return title


def init_db():
    """Initialize the SQLite database for plate tracking."""
    try:
        conn = sqlite3.connect(str(DB_PATH))
        cursor = conn.cursor()
        
        # Create plates table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS plates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plate_id TEXT UNIQUE NOT NULL,
                city TEXT NOT NULL,
                code TEXT NOT NULL,
                plate_number TEXT NOT NULL,
                first_seen_at TEXT NOT NULL,
                last_seen_at TEXT NOT NULL,
                last_telegram_sent_at TEXT,
                total_releases INTEGER DEFAULT 1,
                current_price TEXT,
                seller_name TEXT,
                seller_username TEXT,
                listing_link TEXT,
                created_timestamp REAL NOT NULL,
                updated_timestamp REAL NOT NULL
            )
        ''')
        
        # Create alert_history table (tracks which alerts sent which plates)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS alert_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                alert_id TEXT NOT NULL,
                plate_id TEXT NOT NULL,
                sent_at TEXT NOT NULL,
                created_timestamp REAL NOT NULL,
                FOREIGN KEY (plate_id) REFERENCES plates(plate_id)
            )
        ''')

        # Create daily_listings table (tracks all scraper and OCR raw events)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS daily_listings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                city TEXT NOT NULL,
                plate_code TEXT NOT NULL,
                plate_number TEXT NOT NULL,
                full_plate TEXT NOT NULL,
                digits INTEGER NOT NULL,
                source TEXT NOT NULL,
                price TEXT,
                listing_url TEXT,
                seen_at TEXT NOT NULL,
                raw_data_json TEXT
            )
        ''')

        _ensure_columns(cursor, "daily_listings", {
            "city": "TEXT NOT NULL DEFAULT ''",
            "plate_code": "TEXT NOT NULL DEFAULT ''",
            "plate_number": "TEXT NOT NULL DEFAULT ''",
            "full_plate": "TEXT NOT NULL DEFAULT ''",
            "digits": "INTEGER NOT NULL DEFAULT 0",
            "source": "TEXT NOT NULL DEFAULT 'Website'",
            "price": "TEXT",
            "listing_url": "TEXT",
            "seen_at": "TEXT NOT NULL DEFAULT ''",
            "raw_data_json": "TEXT",
        })
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_daily_listings_seen_at ON daily_listings(seen_at)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_daily_listings_plate_day ON daily_listings(city, plate_code, plate_number, seen_at)")

        # Rule-specific raw events. event_key prevents an unchanged listing from
        # being counted again on every polling pass while retaining price changes,
        # distinct uploads, and different sources as separate events.
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS alert_rule_daily_matches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rule_id TEXT NOT NULL,
                rule_name TEXT NOT NULL,
                event_key TEXT NOT NULL,
                city TEXT NOT NULL,
                plate_code TEXT NOT NULL,
                plate_number TEXT NOT NULL,
                full_plate TEXT NOT NULL,
                digits INTEGER NOT NULL,
                source TEXT NOT NULL,
                price TEXT,
                listing_url TEXT,
                seen_at TEXT NOT NULL,
                raw_data_json TEXT,
                created_at TEXT NOT NULL,
                UNIQUE(rule_id, event_key)
            )
        ''')
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_rule_matches_rule_day ON alert_rule_daily_matches(rule_id, seen_at)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_rule_matches_plate_day ON alert_rule_daily_matches(rule_id, city, plate_code, plate_number, seen_at)")
        
        conn.commit()
        conn.close()
        logger.info("Plate tracking database initialized successfully")
    except Exception as e:
        logger.error(f"Failed to initialize plate tracking database: {e}")
        raise


def get_plate_id(city: str, code: str, plate_number: str) -> str:
    """
    Generate a unique plate ID from city, code, and plate number.
    
    Args:
        city: City name (e.g., "Dubai")
        code: Plate code (e.g., "V")
        plate_number: Plate number (e.g., "51115")
    
    Returns:
        Unique plate identifier (e.g., "dubai_v_51115")
    """
    city_norm = city.strip().lower().replace(' ', '_')
    code_norm = code.strip().lower().replace(' ', '_')
    plate_norm = plate_number.strip().lower().replace(' ', '_')
    return f"{city_norm}_{code_norm}_{plate_norm}"


def track_plate(
    city: str,
    code: str,
    plate_number: str,
    price: str = "",
    seller_name: str = "",
    seller_username: str = "",
    listing_link: str = ""
) -> dict[str, Any]:
    """
    Track or update a plate in the database.
    
    Args:
        city: City name
        code: Plate code
        plate_number: Plate number
        price: Current price (optional)
        seller_name: Seller name (optional)
        seller_username: Seller username (optional)
        listing_link: Listing URL (optional)
    
    Returns:
        Plate tracking record
    """
    plate_id = get_plate_id(city, code, plate_number)
    now = datetime.utcnow()
    now_str = now.strftime('%Y-%m-%d %H:%M:%S')
    now_ts = now.timestamp()
    
    try:
        conn = sqlite3.connect(str(DB_PATH))
        cursor = conn.cursor()
        
        # Check if plate exists
        cursor.execute('SELECT * FROM plates WHERE plate_id = ?', (plate_id,))
        existing = cursor.fetchone()
        
        if existing:
            # Update existing plate (increment release counter)
            cursor.execute('''
                UPDATE plates 
                SET last_seen_at = ?,
                    total_releases = total_releases + 1,
                    current_price = ?,
                    seller_name = ?,
                    seller_username = ?,
                    listing_link = ?,
                    updated_timestamp = ?
                WHERE plate_id = ?
            ''', (now_str, price, seller_name, seller_username, listing_link, now_ts, plate_id))
            
            # Fetch updated record
            cursor.execute('SELECT * FROM plates WHERE plate_id = ?', (plate_id,))
            row = cursor.fetchone()
        else:
            # Create new plate record
            cursor.execute('''
                INSERT INTO plates (
                    plate_id, city, code, plate_number,
                    first_seen_at, last_seen_at,
                    total_releases, current_price,
                    seller_name, seller_username, listing_link,
                    created_timestamp, updated_timestamp
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                plate_id, city, code, plate_number,
                now_str, now_str,
                1, price,
                seller_name, seller_username, listing_link,
                now_ts, now_ts
            ))
            
            # Fetch new record
            cursor.execute('SELECT * FROM plates WHERE plate_id = ?', (plate_id,))
            row = cursor.fetchone()
        
        conn.commit()
        conn.close()
        
        return _row_to_dict(row) if row else None
    except Exception as e:
        logger.error(f"Failed to track plate {plate_id}: {e}")
        raise


def should_send_telegram(
    city: str,
    code: str,
    plate_number: str,
    cooldown_seconds: int = 420,  # 7 minutes default
    alert_id: Optional[str] = None
) -> tuple[bool, dict[str, Any]]:
    """
    Determine if a Telegram alert should be sent for this plate.
    
    Returns:
        (should_send: bool, plate_info: dict)
        - should_send=True if this is a new plate or cooldown has passed
        - plate_info contains tracking data
    """
    plate_id = get_plate_id(city, code, plate_number)
    now = datetime.utcnow()
    
    try:
        conn = sqlite3.connect(str(DB_PATH))
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM plates WHERE plate_id = ?', (plate_id,))
        row = cursor.fetchone()
        plate_info = _row_to_dict(row) if row else None
        
        if not plate_info:
            # New plate - always send
            conn.close()
            return True, {}
        
        # Check if cooldown has passed
        last_sent = plate_info.get('last_telegram_sent_at')
        if not last_sent:
            # Never sent before - send now
            conn.close()
            return True, plate_info
        
        try:
            last_sent_dt = datetime.strptime(last_sent, '%Y-%m-%d %H:%M:%S')
            cooldown_dt = last_sent_dt + timedelta(seconds=cooldown_seconds)
            
            should_send = now >= cooldown_dt
            conn.close()
            return should_send, plate_info
        except Exception as e:
            logger.error(f"Failed to parse last_sent date: {e}")
            conn.close()
            return True, plate_info
            
    except Exception as e:
        logger.error(f"Failed to check should_send_telegram for {plate_id}: {e}")
        return False, {}


def mark_telegram_sent(
    city: str,
    code: str,
    plate_number: str,
    alert_id: Optional[str] = None
) -> None:
    """
    Mark that a Telegram alert was sent for this plate.
    
    Args:
        city: City name
        code: Plate code
        plate_number: Plate number
        alert_id: Alert ID that sent the message (optional, for tracking)
    """
    plate_id = get_plate_id(city, code, plate_number)
    now = datetime.utcnow()
    now_str = now.strftime('%Y-%m-%d %H:%M:%S')
    
    try:
        conn = sqlite3.connect(str(DB_PATH))
        cursor = conn.cursor()
        
        # Update plate
        cursor.execute('''
            UPDATE plates 
            SET last_telegram_sent_at = ?
            WHERE plate_id = ?
        ''', (now_str, plate_id))
        
        # Log to alert history if alert_id provided
        if alert_id:
            cursor.execute('''
                INSERT INTO alert_history (alert_id, plate_id, sent_at, created_timestamp)
                VALUES (?, ?, ?, ?)
            ''', (alert_id, plate_id, now_str, now.timestamp()))
        
        conn.commit()
        conn.close()
        logger.debug(f"Marked Telegram sent for plate {plate_id}")
    except Exception as e:
        logger.error(f"Failed to mark telegram sent for {plate_id}: {e}")
        raise


def get_plate_info(city: str, code: str, plate_number: str) -> Optional[dict[str, Any]]:
    """
    Get tracking information for a plate.
    
    Returns:
        Plate info dict or None if not found
    """
    plate_id = get_plate_id(city, code, plate_number)
    
    try:
        conn = sqlite3.connect(str(DB_PATH))
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM plates WHERE plate_id = ?', (plate_id,))
        row = cursor.fetchone()
        conn.close()
        return _row_to_dict(row) if row else None
    except Exception as e:
        logger.error(f"Failed to get plate info for {plate_id}: {e}")
        return None


def cleanup_old_plates(days: int = 30) -> int:
    """
    Delete plates not seen in the specified number of days.
    
    Args:
        days: Number of days to keep
    
    Returns:
        Number of records deleted
    """
    cutoff_date = (datetime.utcnow() - timedelta(days=days)).strftime('%Y-%m-%d %H:%M:%S')
    
    try:
        conn = sqlite3.connect(str(DB_PATH))
        cursor = conn.cursor()
        cursor.execute('DELETE FROM plates WHERE last_seen_at < ?', (cutoff_date,))
        deleted = cursor.rowcount
        conn.commit()
        conn.close()
        logger.info(f"Cleaned up {deleted} old plates")
        return deleted
    except Exception as e:
        logger.error(f"Failed to cleanup old plates: {e}")
        return 0


def get_plate_stats() -> dict[str, Any]:
    """
    Get statistics about tracked plates.
    
    Returns:
        Dict with counts and info
    """
    try:
        conn = sqlite3.connect(str(DB_PATH))
        cursor = conn.cursor()
        
        cursor.execute('SELECT COUNT(*) FROM plates')
        total_plates = cursor.fetchone()[0]
        
        cursor.execute('SELECT SUM(total_releases) FROM plates')
        total_releases = cursor.fetchone()[0] or 0
        
        cursor.execute('SELECT COUNT(*) FROM plates WHERE last_telegram_sent_at IS NOT NULL')
        alerted_plates = cursor.fetchone()[0]
        
        conn.close()
        return {
            'total_plates_tracked': total_plates,
            'total_releases_seen': total_releases,
            'plates_alerted': alerted_plates
        }
    except Exception as e:
        logger.error(f"Failed to get plate stats: {e}")
        return {}


def _row_to_dict(row: tuple) -> dict[str, Any]:
    """Convert SQLite row to dictionary."""
    if not row:
        return {}
    
    columns = [
        'id', 'plate_id', 'city', 'code', 'plate_number',
        'first_seen_at', 'last_seen_at', 'last_telegram_sent_at',
        'total_releases', 'current_price', 'seller_name',
        'seller_username', 'listing_link', 'created_timestamp', 'updated_timestamp'
    ]
    
    return {col: val for col, val in zip(columns, row)}


def insert_listing_event(
    city: str,
    plate_code: str,
    plate_number: str,
    source: str,
    price: str,
    listing_url: str,
    seen_at: str = None,
    raw_data_json: str = None
) -> None:
    """Insert a raw listing event into the daily_listings table."""
    if not city or not plate_number:
        # Ignore invalid/incomplete entries
        return
        
    city_clean = city.strip().title()
    city_norm_map = {
        "dubai": "Dubai", "دبي": "Dubai",
        "abu dhabi": "Abu Dhabi", "abu-dhabi": "Abu Dhabi", "abudhabi": "Abu Dhabi", "أبوظبي": "Abu Dhabi", "ابوظبي": "Abu Dhabi",
        "sharjah": "Sharjah", "الشارقة": "Sharjah",
        "ajman": "Ajman", "عجمان": "Ajman",
        "ras al khaimah": "Ras Al Khaimah", "ras-al-khaimah": "Ras Al Khaimah", "rak": "Ras Al Khaimah", "رأس الخيمة": "Ras Al Khaimah", "راس الخيمة": "Ras Al Khaimah",
        "umm al quwain": "Umm Al Quwain", "umm-al-quwain": "Umm Al Quwain", "umm al qaiwain": "Umm Al Quwain", "أم القيوين": "Umm Al Quwain", "ام القيوين": "Umm Al Quwain",
        "fujairah": "Fujairah", "الفجيرة": "Fujairah"
    }
    norm_key = city_clean.lower()
    city_clean = city_norm_map.get(norm_key, city_clean)
    
    code_clean = plate_code.strip().upper() if plate_code else ""
    if code_clean == "?" or code_clean.lower() == "any code":
        code_clean = ""
        
    num_clean = plate_number.strip()
    
    if code_clean:
        full_plate = f"{city_clean} {code_clean} {num_clean}"
    else:
        full_plate = f"{city_clean} {num_clean}"
        
    digits = len([c for c in num_clean if c.isdigit()])
    
    source_clean = source.strip()
    if not source_clean:
        source_clean = "Website"
    elif "ocr" in source_clean.lower():
        source_clean = "OCR Instagram"
    elif "instagram" in source_clean.lower():
        source_clean = "Instagram"
    else:
        source_clean = "Website"
        
    price_clean = _price_for_report(price)
        
    if not seen_at:
        seen_at = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    elif isinstance(seen_at, datetime):
        seen_at = seen_at.strftime('%Y-%m-%d %H:%M:%S')
    else:
        seen_at = str(seen_at).strip()

    if raw_data_json is not None and not isinstance(raw_data_json, str):
        raw_data_json = json.dumps(raw_data_json, ensure_ascii=False, default=str)
        
    try:
        conn = sqlite3.connect(str(DB_PATH))
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO daily_listings (
                city, plate_code, plate_number, full_plate, digits,
                source, price, listing_url, seen_at, raw_data_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            city_clean, code_clean, num_clean, full_plate, digits,
            source_clean, price_clean, listing_url, seen_at, raw_data_json
        ))
        conn.commit()
        conn.close()
        logger.debug(f"Logged daily listing event for {full_plate}")
    except Exception as e:
        logger.error(f"Failed to log listing event for {full_plate}: {e}")


def _normalize_event_source(value: Any) -> str:
    source = str(value or "").strip()
    lowered = source.lower()
    if "ocr" in lowered:
        return "OCR Instagram"
    if "instagram" in lowered:
        return "Instagram"
    if not source or any(term in lowered for term in ("xplate", "website", "web")):
        return "Website"
    return source


def _rule_event_key(rule_id: str, listing: dict[str, Any], source: str, price: str) -> str:
    listing_url = str(
        listing.get("listing_link")
        or listing.get("listing_url")
        or listing.get("post_url")
        or ""
    ).strip()
    listing_id = str(listing.get("listing_id") or listing.get("id") or "").strip()
    posted_marker = "|".join(str(listing.get(field) or "").strip() for field in (
        "uploaded_date", "uploaded_time", "posted_at", "posted_time", "post_id"
    ))
    identity = {
        "rule_id": rule_id,
        "listing_url": listing_url,
        "listing_id": listing_id,
        "city": str(listing.get("city") or "").strip().lower(),
        "code": str(listing.get("code") or listing.get("plate_code") or "").strip().upper(),
        "plate_number": str(listing.get("plate_number") or "").strip(),
        "source": source,
        "price": price,
        "posted_marker": posted_marker,
    }
    encoded = json.dumps(identity, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def insert_alert_rule_match(
    rule_id: str,
    rule_name: str,
    listing: dict[str, Any],
    seen_at: str | datetime | None = None,
) -> bool:
    """Persist one rule-matching listing event, returning True when newly inserted."""
    city = str(listing.get("city") or "").strip()
    plate_number = str(listing.get("plate_number") or "").strip()
    if not rule_id or not city or not plate_number:
        logger.warning("Rule match skipped because rule ID, city, or plate number is missing")
        return False

    code = str(listing.get("code") or listing.get("plate_code") or "").strip().upper()
    if code in {"?", "ANY CODE"}:
        code = ""
    city = city.title()
    city_aliases = {
        "Rak": "Ras Al Khaimah",
        "Ras-Al-Khaimah": "Ras Al Khaimah",
        "Uaq": "Umm Al Quwain",
        "Umm-Al-Quwain": "Umm Al Quwain",
        "Abudhabi": "Abu Dhabi",
        "Abu-Dhabi": "Abu Dhabi",
    }
    city = city_aliases.get(city, city)
    full_plate = " ".join(part for part in (city, code, plate_number) if part)
    source = _normalize_event_source(
        listing.get("source") or listing.get("listing_source") or listing.get("platform")
    )
    price = _price_for_report(listing.get("price"))
    listing_url = str(
        listing.get("listing_link")
        or listing.get("listing_url")
        or listing.get("post_url")
        or ""
    ).strip()
    if isinstance(seen_at, datetime):
        seen_at_text = seen_at.astimezone(ZoneInfo("Asia/Dubai")).strftime("%Y-%m-%d %H:%M:%S")
    elif seen_at:
        seen_at_text = str(seen_at).strip()
    else:
        seen_at_text = datetime.now(ZoneInfo("Asia/Dubai")).strftime("%Y-%m-%d %H:%M:%S")
    raw_data_json = json.dumps(listing, ensure_ascii=False, default=str)
    event_key = _rule_event_key(rule_id, listing, source, price)

    conn = sqlite3.connect(str(DB_PATH))
    try:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT OR IGNORE INTO alert_rule_daily_matches (
                rule_id, rule_name, event_key, city, plate_code, plate_number,
                full_plate, digits, source, price, listing_url, seen_at,
                raw_data_json, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            str(rule_id), str(rule_name or rule_id), event_key, city, code,
            plate_number, full_plate, sum(char.isdigit() for char in plate_number),
            source, price, listing_url, seen_at_text, raw_data_json,
            datetime.now(ZoneInfo("Asia/Dubai")).isoformat(),
        ))
        inserted = cursor.rowcount > 0
        conn.commit()
        if inserted:
            logger.info("Stored daily rule match: rule=%s plate=%s", rule_id, full_plate)
        return inserted
    finally:
        conn.close()


def store_alert_rule_matches(rule_id: str, rule_name: str, listings: list[dict[str, Any]]) -> int:
    """Store all filtered matches from one scan and return the number of new events."""
    inserted = 0
    for listing in listings:
        try:
            inserted += int(insert_alert_rule_match(rule_id, rule_name, listing))
        except Exception:
            logger.exception("Failed storing daily rule match: rule=%s", rule_id)
    return inserted


def get_alert_rule_matches_by_date(rule_id: str, date_str: str) -> list[dict[str, Any]]:
    _parse_report_date(date_str)
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute('''
            SELECT id, rule_id, rule_name, city, plate_code, plate_number,
                   full_plate, digits, source, price, listing_url, seen_at
            FROM alert_rule_daily_matches
            WHERE rule_id = ? AND substr(seen_at, 1, 10) = ?
            ORDER BY seen_at ASC, id ASC
        ''', (str(rule_id), date_str)).fetchall()
        return [dict(row) for row in rows]
    finally:
        conn.close()


def generate_notes(sources: list[str], prices: list[str], times_uploaded: int, links: list[str] | None = None) -> str:
    """Generate notes about the listings of a plate on a given day."""
    notes = []
    
    has_ocr = any("ocr" in s.lower() for s in sources)
    
    unique_sources = set(sources)
    if len(unique_sources) > 1:
        notes.append("multiple sources")
        
    if times_uploaded > 1:
        notes.append(f"same plate uploaded {times_uploaded} times")

    if links and len(set(links)) < len(links):
        notes.append("duplicate listing link")
        
    valid_prices = [p for p in prices if not _is_missing_price(p)]
    missing_prices = [p for p in prices if _is_missing_price(p)]
    if not valid_prices:
        notes.append("missing price")
    else:
        unique_prices = set(valid_prices)
        if len(unique_prices) > 1:
            notes.append("multiple prices")
        if missing_prices:
            notes.append("missing price captured")
            
    if has_ocr:
        notes.append("OCR result")
            
    if not notes:
        notes.append("New listing")
        
    result_notes = ", ".join(notes)
    if result_notes:
        result_notes = result_notes[0].upper() + result_notes[1:]
    return result_notes


def get_daily_listings_by_date(date_str: str) -> list[dict[str, Any]]:
    """Get raw daily listings from database by date YYYY-MM-DD."""
    try:
        _parse_report_date(date_str)
        conn = sqlite3.connect(str(DB_PATH))
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, city, plate_code, plate_number, full_plate, digits, source, price, listing_url, seen_at
            FROM daily_listings
            WHERE substr(seen_at, 1, 10) = ?
            ORDER BY seen_at ASC, id ASC
        ''', (date_str,))
        rows = cursor.fetchall()
        conn.close()
        
        result = []
        for r in rows:
            result.append({
                "id": r[0],
                "city": r[1],
                "plate_code": r[2],
                "plate_number": r[3],
                "full_plate": r[4],
                "digits": r[5],
                "source": r[6],
                "price": r[7],
                "listing_url": r[8],
                "seen_at": r[9]
            })
        return result
    except Exception as e:
        logger.error(f"Failed to fetch daily listings for date {date_str}: {e}")
        return []


def _aggregate_listing_events(events: list[dict[str, Any]], date_str: str) -> dict[str, Any]:
    """Aggregate raw listing events into grouped city sheets and summary metrics."""
    grouped = {}
    for ev in events:
        key = (ev["city"].strip().title(), ev["plate_code"].strip().upper(), ev["plate_number"].strip())
        if key not in grouped:
            grouped[key] = []
        grouped[key].append(ev)
        
    city_sheets = {}
    
    total_unique_plates = len(grouped)
    total_listing_events = len(events)
    total_repeated_plates = 0
    total_sources = set()
    total_cities = set()
    
    plates_per_city = {}
    events_per_city = {}
    most_repeated_plate = "N/A"
    max_repeats = 0
    missing_price_count = 0
    instagram_ocr_count = 0
    instagram_count = 0
    website_count = 0
    
    for key, ev_list in grouped.items():
        city, code, num = key
        total_cities.add(city)
        plates_per_city[city] = plates_per_city.get(city, 0) + 1
        events_per_city[city] = events_per_city.get(city, 0) + len(ev_list)
        
        times_uploaded = len(ev_list)
        if times_uploaded > 1:
            total_repeated_plates += 1
            
        if times_uploaded > max_repeats:
            max_repeats = times_uploaded
            if code:
                most_repeated_plate = f"{city} {code} {num} ({times_uploaded} times)"
            else:
                most_repeated_plate = f"{city} {num} ({times_uploaded} times)"
                
        sources = []
        prices = []
        links = []
        
        for ev in ev_list:
            total_sources.add(ev["source"])
            sources.append(ev["source"])
            prices.append(_price_for_report(ev["price"]))
            if ev["listing_url"]:
                links.append(ev["listing_url"])
                
        unique_sources = []
        for s in sources:
            if s not in unique_sources:
                unique_sources.append(s)
        source_str = ", ".join(unique_sources)
        
        latest_price = "N/A"
        for p in reversed(prices):
            if p != "N/A":
                latest_price = p
                break
                
        if any(_is_missing_price(p) for p in prices):
            missing_price_count += 1
            
        all_prices_seen = ", ".join(prices)
        
        unique_links = []
        for l in links:
            if l not in unique_links:
                unique_links.append(l)
        links_str = ", ".join(unique_links)
        
        if any("ocr" in s.lower() for s in unique_sources):
            instagram_ocr_count += 1
        elif any("instagram" in s.lower() for s in unique_sources):
            instagram_count += 1
        if any("website" in s.lower() for s in unique_sources):
            website_count += 1
            
        notes_str = generate_notes(unique_sources, prices, times_uploaded, links)
        
        if code:
            full_plate = f"{city} {code} {num}"
        else:
            full_plate = f"{city} {num}"
            
        digits = len([c for c in num if c.isdigit()])
        
        city_sheet_name = city
        if city_sheet_name == "Ras Al Khaimah":
            city_sheet_name = "RAK"
        elif city_sheet_name == "Umm Al Quwain":
            city_sheet_name = "UAQ"
            
        if city_sheet_name not in city_sheets:
            city_sheets[city_sheet_name] = []
            
        city_sheets[city_sheet_name].append({
            "Full Plate": full_plate,
            "Digits": digits,
            "Source": source_str,
            "Times Uploaded Today": times_uploaded,
            "Price": latest_price,
            "All Prices Seen": all_prices_seen,
            "Listing Links": links_str,
            "Notes": notes_str
        })
        
    return {
        "summary": {
            "date": date_str,
            "total_unique_plates": total_unique_plates,
            "total_listing_events": total_listing_events,
            "total_repeated_plates": total_repeated_plates,
            "total_sources": len(total_sources),
            "total_cities": len(total_cities),
            "plates_per_city": plates_per_city,
            "events_per_city": events_per_city,
            "most_repeated_plate": most_repeated_plate,
            "missing_price_count": missing_price_count,
            "instagram_ocr_count": instagram_ocr_count,
            "instagram_count": instagram_count,
            "website_count": website_count
        },
        "city_sheets": city_sheets
    }


def aggregate_daily_report(date_str: str) -> dict[str, Any]:
    """Aggregate raw daily listing events into a daily report structure."""
    return _aggregate_listing_events(get_daily_listings_by_date(date_str), date_str)


def aggregate_daily_rule_report(rule_id: str, date_str: str) -> dict[str, Any]:
    """Aggregate daily listing events that were captured for one saved rule."""
    return _aggregate_listing_events(get_alert_rule_matches_by_date(rule_id, date_str), date_str)


def generate_daily_excel_report(date_str: str) -> str:
    """Generate the Excel Daily Report workbook."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    _parse_report_date(date_str)
    
    report_data = aggregate_daily_report(date_str)
    summary = report_data["summary"]
    city_sheets = report_data["city_sheets"]
    
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    font_title = Font(name="Segoe UI", size=16, bold=True, color="1F4E79")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=11)
    font_body_bold = Font(name="Segoe UI", size=11, bold=True)
    
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_accent = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    thin_side = Side(border_style="thin", color="D3D3D3")
    border_thin = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def style_header_row(sheet, row_number: int, last_column: int) -> None:
        for col_idx in range(1, last_column + 1):
            cell = sheet.cell(row=row_number, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def auto_width(sheet, max_width: int = 55) -> None:
        for col in sheet.columns:
            max_len = 0
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            column_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[column_letter].width = min(max(max_len + 3, 12), max_width)
    
    ws_summary.freeze_panes = "A2"
    ws_summary.cell(row=1, column=1, value="Xplate Daily Report").font = font_title
    ws_summary.cell(row=2, column=1, value=f"Report date: {date_str}").font = Font(name="Segoe UI", size=11, italic=True)
    
    if not city_sheets:
        ws_summary.cell(row=4, column=1, value="No data found for the selected date.").font = font_body_bold
        ws_summary.cell(row=4, column=1).fill = fill_accent
        ws_summary.cell(row=6, column=1, value="Metric")
        ws_summary.cell(row=6, column=2, value="Value")
        style_header_row(ws_summary, 6, 2)
        for row_idx, (metric, value) in enumerate([
            ("Total unique plates uploaded", 0),
            ("Total listing events", 0),
            ("Total repeated plates", 0),
            ("Total sources", 0),
            ("Total cities included", 0),
        ], start=7):
            ws_summary.cell(row=row_idx, column=1, value=metric).font = font_body
            ws_summary.cell(row=row_idx, column=2, value=value).font = font_body_bold
            ws_summary.cell(row=row_idx, column=1).border = border_thin
            ws_summary.cell(row=row_idx, column=2).border = border_thin
    else:
        kpis = [
            ("Total unique plates uploaded", summary["total_unique_plates"]),
            ("Total listing events", summary["total_listing_events"]),
            ("Total repeated plates", summary["total_repeated_plates"]),
            ("Total sources", summary["total_sources"]),
            ("Total cities included", summary["total_cities"]),
            ("Most repeated plate of the day", summary["most_repeated_plate"]),
            ("Number of plates with missing price", summary["missing_price_count"]),
            ("Number of plates found from Instagram/OCR", summary["instagram_count"] + summary["instagram_ocr_count"]),
            ("Number of plates found from websites", summary["website_count"]),
        ]
        
        ws_summary.cell(row=4, column=1, value="Summary").font = font_section
        
        ws_summary.cell(row=5, column=1, value="Metric")
        ws_summary.cell(row=5, column=2, value="Value")
        style_header_row(ws_summary, 5, 2)
        
        curr_row = 6
        for name, val in kpis:
            c_name = ws_summary.cell(row=curr_row, column=1, value=name)
            c_name.font = font_body
            c_name.border = border_thin
            
            c_val = ws_summary.cell(row=curr_row, column=2, value=val)
            c_val.font = font_body_bold
            c_val.border = border_thin
            if name in {"Total unique plates uploaded", "Total listing events"}:
                c_val.fill = fill_accent
            curr_row += 1
            
        breakdown_header_row = curr_row + 2
        ws_summary.cell(row=breakdown_header_row, column=1, value="Breakdown by Emirate").font = font_section
        
        table_header_row = breakdown_header_row + 1
        ws_summary.cell(row=table_header_row, column=1, value="Emirate")
        ws_summary.cell(row=table_header_row, column=2, value="Unique Plates")
        ws_summary.cell(row=table_header_row, column=3, value="Listing Events")
        style_header_row(ws_summary, table_header_row, 3)
        
        city_row = table_header_row + 1
        for city, count in sorted(summary["plates_per_city"].items()):
            city_display = city
            if city_display == "Ras Al Khaimah":
                city_display = "RAK"
            elif city_display == "Umm Al Quwain":
                city_display = "UAQ"
                
            c_city = ws_summary.cell(row=city_row, column=1, value=city_display)
            c_city.font = font_body
            c_city.border = border_thin
            
            c_unique = ws_summary.cell(row=city_row, column=2, value=count)
            c_unique.font = font_body
            c_unique.border = border_thin
            c_unique.alignment = Alignment(horizontal="right")
            
            c_events = ws_summary.cell(row=city_row, column=3, value=summary["events_per_city"].get(city, 0))
            c_events.font = font_body
            c_events.border = border_thin
            c_events.alignment = Alignment(horizontal="right")
            city_row += 1
            
    auto_width(ws_summary, max_width=70)
        
    headers = [
        "Full Plate", "Digits", "Source", "Times Uploaded Today",
        "Price", "All Prices Seen", "Listing Links", "Notes"
    ]
    
    used_sheet_names = {"Summary"}
    for city, rows in sorted(city_sheets.items()):
        sheet_title = _safe_sheet_title(city, used_sheet_names)
        ws_city = wb.create_sheet(title=sheet_title)
        ws_city.freeze_panes = "A2"
        
        for col_idx, h in enumerate(headers, start=1):
            cell = ws_city.cell(row=1, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="left" if h not in ["Digits", "Times Uploaded Today", "Price"] else "right", vertical="center")
            cell.border = border_thin
            
        sorted_rows = sorted(rows, key=lambda item: (-int(item.get("Times Uploaded Today") or 0), str(item.get("Full Plate") or "")))
        for row_idx, r_data in enumerate(sorted_rows, start=2):
            for col_idx, h in enumerate(headers, start=1):
                val = r_data[h]
                cell = ws_city.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_body
                cell.border = border_thin
                
                if h in ["Digits", "Times Uploaded Today"]:
                    cell.alignment = Alignment(horizontal="right")
                    try:
                        cell.value = int(val)
                    except (ValueError, TypeError):
                        cell.value = val
                elif h == "Price":
                    cell.alignment = Alignment(horizontal="right")
                    if val != "N/A":
                        clean_num = re.sub(r"[^\d.]", "", val)
                        if clean_num:
                            try:
                                cell.value = float(clean_num) if "." in clean_num else int(clean_num)
                                cell.number_format = '#,##0'
                            except ValueError:
                                cell.value = val
                elif h == "Listing Links":
                    if val and val != "N/A":
                        links_list = [l.strip() for l in val.split(",") if l.strip()]
                        if links_list:
                            link_labels = [f"Link {i+1}" for i in range(len(links_list))]
                            cell.value = ", ".join(link_labels)
                            cell.hyperlink = links_list[0]
                            cell.font = Font(name="Segoe UI", size=11, color="0563C1", underline="single")
                        else:
                            cell.value = "N/A"
                    else:
                        cell.value = "N/A"
                else:
                    cell.alignment = Alignment(horizontal="left")
                    
        ws_city.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(sorted_rows) + 1}"
        auto_width(ws_city, max_width=45)
            
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    
    file_name = daily_report_filename(date_str)
    file_path = REPORTS_DIR / file_name
    
    wb.save(str(file_path))
    return str(file_path.resolve())


def _find_saved_rule(rule_id: str) -> dict[str, Any]:
    from .storage import get_alerts

    rule = next((item for item in get_alerts() if str(item.get("id") or "") == str(rule_id)), None)
    if not rule:
        raise KeyError(f"Alert rule not found: {rule_id}")
    return rule


def _rule_filters_for_report(rule: dict[str, Any]) -> dict[str, Any]:
    filter_keys = (
        "cities", "city", "code", "plate_number", "search_mode", "number_formats",
        "number_format", "price_min", "price_max", "contains", "starts_with",
        "ends_with", "send_all_new_plates", "include_featured_listings",
        "include_sold_listings",
    )
    filters: dict[str, Any] = {}
    for key in filter_keys:
        value = rule.get(key)
        if value not in (None, "", [], False):
            filters[key] = value
    return filters


def generate_daily_rule_excel_report(
    rule_id: str,
    date_str: str,
    telegram_sent_status: str = "Not requested",
) -> Path:
    """Generate a professional daily workbook containing only one rule's matches."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    _parse_report_date(date_str)
    rule = _find_saved_rule(rule_id)
    enabled_value = rule.get("enabled")
    if enabled_value is not True and str(enabled_value or "").strip().lower() not in {"true", "1", "yes", "on"}:
        raise ValueError("Disabled rules do not generate or send daily reports.")
    report_data = aggregate_daily_rule_report(rule_id, date_str)
    summary = report_data["summary"]
    city_sheets = report_data["city_sheets"]
    rule_name = str(rule.get("name") or rule_id)
    rule_cities = rule.get("cities") or ([rule.get("city")] if rule.get("city") else [])
    filters_text = json.dumps(_rule_filters_for_report(rule), ensure_ascii=False, sort_keys=True)

    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1F4E79")
    section_font = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Segoe UI", size=11)
    bold_font = Font(name="Segoe UI", size=11, bold=True)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    accent_fill = PatternFill("solid", fgColor="DDEBF7")
    border = Border(*(Side(style="thin", color="D3D3D3") for _ in range(4)))

    def style_header(sheet, row: int, last_column: int) -> None:
        for column in range(1, last_column + 1):
            cell = sheet.cell(row=row, column=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

    def auto_width(sheet, max_width: int = 55) -> None:
        for column_cells in sheet.columns:
            longest = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(longest + 3, 12), max_width)

    summary_sheet.freeze_panes = "A2"
    summary_sheet["A1"] = "Xplate Scout Daily Rule Report"
    summary_sheet["A1"].font = title_font
    summary_sheet["A2"] = f"{rule_name} — {date_str}"
    summary_sheet["A2"].font = Font(name="Segoe UI", size=11, italic=True)
    if not city_sheets:
        summary_sheet["A4"] = "No data found for this saved rule on the selected date."
        summary_sheet["A4"].font = bold_font
        summary_sheet["A4"].fill = accent_fill

    summary_sheet["A6"] = "Report Details"
    summary_sheet["A6"].font = section_font
    summary_sheet["A7"] = "Metric"
    summary_sheet["B7"] = "Value"
    style_header(summary_sheet, 7, 2)
    detail_rows = [
        ("Report type", "Daily Saved Rule Report"),
        ("Rule name", rule_name),
        ("Rule ID", str(rule_id)),
        ("Report date", date_str),
        ("Cities included", ", ".join(str(city) for city in rule_cities if city) or "All cities"),
        ("Filters used", filters_text or "Broad rule (no filters)"),
        ("Total unique plates", summary["total_unique_plates"]),
        ("Total listing events", summary["total_listing_events"]),
        ("Total repeated plates", summary["total_repeated_plates"]),
        ("Total sources", summary["total_sources"]),
        ("Most repeated plate", summary["most_repeated_plate"]),
        ("Number of plates with missing price", summary["missing_price_count"]),
        ("Number of Instagram/OCR matches", summary["instagram_count"] + summary["instagram_ocr_count"]),
        ("Number of website matches", summary["website_count"]),
        ("Telegram sent status", telegram_sent_status),
    ]
    for row_number, (label, value) in enumerate(detail_rows, start=8):
        summary_sheet.cell(row=row_number, column=1, value=label)
        summary_sheet.cell(row=row_number, column=2, value=value)
        for column in (1, 2):
            summary_sheet.cell(row=row_number, column=column).border = border
            summary_sheet.cell(row=row_number, column=column).font = bold_font if column == 2 else body_font
        if label in {"Total unique plates", "Total listing events"}:
            summary_sheet.cell(row=row_number, column=2).fill = accent_fill

    breakdown_row = 8 + len(detail_rows) + 2
    summary_sheet.cell(row=breakdown_row, column=1, value="Breakdown by Emirate").font = section_font
    summary_sheet.cell(row=breakdown_row + 1, column=1, value="Emirate")
    summary_sheet.cell(row=breakdown_row + 1, column=2, value="Unique Plates")
    summary_sheet.cell(row=breakdown_row + 1, column=3, value="Listing Events")
    style_header(summary_sheet, breakdown_row + 1, 3)
    if summary["plates_per_city"]:
        city_breakdowns = sorted(summary["plates_per_city"].items())
    else:
        city_breakdowns = [("No data", 0)]
    for row_number, (city, unique_count) in enumerate(city_breakdowns, start=breakdown_row + 2):
        display_city = {"Ras Al Khaimah": "RAK", "Umm Al Quwain": "UAQ"}.get(city, city)
        values = (display_city, unique_count, summary["events_per_city"].get(city, 0))
        for column, value in enumerate(values, start=1):
            cell = summary_sheet.cell(row=row_number, column=column, value=value)
            cell.border = border
            cell.font = body_font
    auto_width(summary_sheet, 80)

    headers = [
        "Full Plate", "Digits", "Source", "Times Uploaded Today",
        "Price", "All Prices Seen", "Listing Links", "Notes",
    ]
    used_sheet_names = {"Summary"}
    for city, rows in sorted(city_sheets.items()):
        city_sheet = workbook.create_sheet(_safe_sheet_title(city, used_sheet_names))
        city_sheet.freeze_panes = "A2"
        for column, header in enumerate(headers, start=1):
            cell = city_sheet.cell(row=1, column=column, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if header in {"Digits", "Times Uploaded Today", "Price"} else "left")
        sorted_rows = sorted(rows, key=lambda item: (-int(item["Times Uploaded Today"]), item["Full Plate"]))
        for row_number, row_data in enumerate(sorted_rows, start=2):
            for column, header in enumerate(headers, start=1):
                value = row_data[header]
                cell = city_sheet.cell(row=row_number, column=column, value=value)
                cell.font = body_font
                cell.border = border
                if header in {"Digits", "Times Uploaded Today"}:
                    cell.value = int(value)
                    cell.alignment = Alignment(horizontal="right")
                elif header == "Price" and value != "N/A":
                    numeric_price = re.sub(r"[^\d.]", "", str(value))
                    if numeric_price:
                        cell.value = float(numeric_price) if "." in numeric_price else int(numeric_price)
                        cell.number_format = "#,##0"
                        cell.alignment = Alignment(horizontal="right")
                elif header == "Listing Links":
                    links = [link.strip() for link in str(value or "").split(",") if link.strip()]
                    if links:
                        cell.value = ", ".join(f"Link {index}" for index in range(1, len(links) + 1))
                        cell.hyperlink = links[0]
                        cell.font = Font(name="Segoe UI", size=11, color="0563C1", underline="single")
                    else:
                        cell.value = "N/A"
        city_sheet.auto_filter.ref = f"A1:H{len(sorted_rows) + 1}"
        auto_width(city_sheet, 48)

    safe_rule_id = re.sub(r"[^A-Za-z0-9_-]+", "_", str(rule_id)).strip("_") or "rule"
    rule_reports_dir = RULE_REPORTS_DIR / safe_rule_id
    rule_reports_dir.mkdir(parents=True, exist_ok=True)
    file_path = (rule_reports_dir / daily_report_filename(date_str)).resolve()
    workbook.save(file_path)
    logger.info("Daily rule Excel generated: rule=%s date=%s path=%s", rule_id, date_str, file_path)
    return file_path


# Initialize database on module import
try:
    init_db()
except Exception as e:
    logger.warning(f"Database initialization skipped: {e}")
