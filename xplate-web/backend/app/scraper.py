import argparse
import json
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Callable
from urllib.parse import parse_qs, quote_plus, unquote, urlencode, urljoin, urlparse, urlunparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


BASE_URL = "https://xplate.com/en/numbers/license-plates"

CITY_TO_XPLATE_PARAM = {
    "dubai": "dubai",
    "دبي": "dubai",
    "abu dhabi": "abu+dhabi",
    "abu-dhabi": "abu+dhabi",
    "abudhabi": "abu+dhabi",
    "أبوظبي": "abu+dhabi",
    "ابوظبي": "abu+dhabi",
    "sharjah": "sharjah",
    "الشارقة": "sharjah",
    "ajman": "ajman",
    "عجمان": "ajman",
    "ras al khaimah": "ras+al+khaimah",
    "ras-al-khaimah": "ras+al+khaimah",
    "rak": "ras+al+khaimah",
    "r.a.k": "ras+al+khaimah",
    "r a k": "ras+al+khaimah",
    "رأس الخيمة": "ras+al+khaimah",
    "راس الخيمة": "ras+al+khaimah",
    "umm al quwain": "umm+al+quwain",
    "umm-al-quwain": "umm+al+quwain",
    "umm al qaiwain": "umm+al+quwain",
    "أم القيوين": "umm+al+quwain",
    "ام القيوين": "umm+al+quwain",
    "fujairah": "fujairah",
    "الفجيرة": "fujairah",
}

CITIES = [
    "abu dhabi",
    "dubai",
    "sharjah",
    "ajman",
    "umm al quwain",
    "ras al khaimah",
    "fujairah",
]

RESULT_COLUMNS = [
    "city",
    "plate_number",
    "code",
    "price",
    "seller_name",
    "seller_username",
    "phone_number",
    "uploaded_date",
    "uploaded_time",
    "age_text",
    "deal_rank",
    "featured",
    "sold",
    "listing_type",
    "views",
    "listing_id",
    "listing_link",
    "seller_link",
    "source_url",
]

NUMBER_FORMAT_CATALOG = [
    {"id": "repeat_2", "label": "Contains digit repeated 2 times", "group": "General", "repeat": 2},
    {"id": "repeat_3", "label": "Contains digit repeated 3 times", "group": "General", "repeat": 3},
    {"id": "repeat_4", "label": "Contains digit repeated 4 times", "group": "General", "repeat": 4},
    {"id": "x???x_5", "label": "x???x (5 Digits)", "group": "5 Digits", "pattern": "x???x", "length": 5},
    {"id": "xyzyx_5", "label": "xyzyx (5 Digits)", "group": "5 Digits", "pattern": "xyzyx", "length": 5},
    {"id": "xxyxx_5", "label": "xxyxx (5 Digits)", "group": "5 Digits", "pattern": "xxyxx", "length": 5},
    {"id": "?xxx?_5", "label": "?xxx? (5 Digits)", "group": "5 Digits", "pattern": "?xxx?", "length": 5},
    {"id": "xyxyx_5", "label": "xyxyx (5 Digits)", "group": "5 Digits", "pattern": "xyxyx", "length": 5},
    {"id": "xyyyx_5", "label": "xyyyx (5 Digits)", "group": "5 Digits", "pattern": "xyyyx", "length": 5},
    {"id": "??xx?_5", "label": "??xx? (5 Digits)", "group": "5 Digits", "pattern": "??xx?", "length": 5},
    {"id": "xxx??_5", "label": "xxx?? (5 Digits)", "group": "5 Digits", "pattern": "xxx??", "length": 5},
    {"id": "xyyyy_5", "label": "xyyyy (5 Digits)", "group": "5 Digits", "pattern": "xyyyy", "length": 5},
    {"id": "xxyyy_5", "label": "xxyyy (5 Digits)", "group": "5 Digits", "pattern": "xxyyy", "length": 5},
    {"id": "xxxyy_5", "label": "xxxyy (5 Digits)", "group": "5 Digits", "pattern": "xxxyy", "length": 5},
    {"id": "xxxyx_5", "label": "xxxyx (5 Digits)", "group": "5 Digits", "pattern": "xxxyx", "length": 5},
    {"id": "xyxxx_5", "label": "xyxxx (5 Digits)", "group": "5 Digits", "pattern": "xyxxx", "length": 5},
    {"id": "xxxxy_5", "label": "xxxxy (5 Digits)", "group": "5 Digits", "pattern": "xxxxy", "length": 5},
    {"id": "xyxxy_5", "label": "xyxxy (5 Digits)", "group": "5 Digits", "pattern": "xyxxy", "length": 5},
    {"id": "xxxxx_5", "label": "xxxxx (5 Digits)", "group": "5 Digits", "pattern": "xxxxx", "length": 5},
    {"id": "x??x_4", "label": "x??x (4 Digits)", "group": "4 Digits", "pattern": "x??x", "length": 4},
    {"id": "xyyx_4", "label": "xyyx (4 Digits)", "group": "4 Digits", "pattern": "xyyx", "length": 4},
    {"id": "xyxy_4", "label": "xyxy (4 Digits)", "group": "4 Digits", "pattern": "xyxy", "length": 4},
    {"id": "?xx?_4", "label": "?xx? (4 Digits)", "group": "4 Digits", "pattern": "?xx?", "length": 4},
    {"id": "xxxy_4", "label": "xxxy (4 Digits)", "group": "4 Digits", "pattern": "xxxy", "length": 4},
    {"id": "xyyy_4", "label": "xyyy (4 Digits)", "group": "4 Digits", "pattern": "xyyy", "length": 4},
    {"id": "xyxx_4", "label": "xyxx (4 Digits)", "group": "4 Digits", "pattern": "xyxx", "length": 4},
    {"id": "xxyx_4", "label": "xxyx (4 Digits)", "group": "4 Digits", "pattern": "xxyx", "length": 4},
    {"id": "xxyy_4", "label": "xxyy (4 Digits)", "group": "4 Digits", "pattern": "xxyy", "length": 4},
    {"id": "xxxx_4", "label": "xxxx (4 Digits)", "group": "4 Digits", "pattern": "xxxx", "length": 4},
    {"id": "xyx_3", "label": "xyx (3 Digits)", "group": "3 Digits", "pattern": "xyx", "length": 3},
    {"id": "xyz_3", "label": "xyz (3 Digits)", "group": "3 Digits", "pattern": "xyz", "length": 3},
    {"id": "xyy_3", "label": "xyy (3 Digits)", "group": "3 Digits", "pattern": "xyy", "length": 3},
    {"id": "xxy_3", "label": "xxy (3 Digits)", "group": "3 Digits", "pattern": "xxy", "length": 3},
    {"id": "xxx_3", "label": "xxx (3 Digits)", "group": "3 Digits", "pattern": "xxx", "length": 3},
]

NUMBER_FORMAT_OPTIONS = [
    "Any format",
    *[item["label"] for item in NUMBER_FORMAT_CATALOG],
]

FORMAT_BY_ID = {item["id"]: item for item in NUMBER_FORMAT_CATALOG}
FORMAT_BY_LABEL = {item["label"]: item for item in NUMBER_FORMAT_CATALOG}
FORMAT_LABEL_BY_ID = {item["id"]: item["label"] for item in NUMBER_FORMAT_CATALOG}
FORMAT_TO_URL_VALUE = {"Any format": "", **{item["label"]: item.get("pattern", "") for item in NUMBER_FORMAT_CATALOG}}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}


def debug_print(message: str, debug_callback: Callable[[str], None] | None = None) -> None:
    print(message)
    if debug_callback:
        debug_callback(message)


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def city_to_xplate_param(city: str | None) -> str:
    text = clean_text(str(city or "").lower().replace("-", " ").replace("_", " "))
    if not text or text in {"all", "all cities"}:
        return ""
    compact = text.replace(" ", "")
    return (
        CITY_TO_XPLATE_PARAM.get(text)
        or CITY_TO_XPLATE_PARAM.get(text.replace(" ", "-"))
        or CITY_TO_XPLATE_PARAM.get(compact)
        or text.replace(" ", "+")
    )


def city_to_url_value(city: str | None) -> str:
    return city_to_xplate_param(city)


def normalize_search_mode(search_mode: str) -> str:
    mode = (search_mode or "contains").strip().lower().replace("_", " ")
    if mode in {"starts", "start", "starts with", "starts-with"}:
        return "starts with"
    if mode in {"ends", "end", "ends with", "ends-with"}:
        return "ends with"
    if mode in {"exact", "exact match", "exact-match"}:
        return "exact match"
    return "contains"


def normalize_number_formats(value: str | list[str] | tuple[str, ...] | None, fallback: str | None = None) -> list[str]:
    raw_values: list[str] = []
    if isinstance(value, (list, tuple)):
        raw_values.extend(str(item or "").strip() for item in value)
    elif value is not None:
        raw_values.append(str(value or "").strip())
    if not raw_values and fallback:
        raw_values.append(str(fallback or "").strip())

    normalized: list[str] = []
    for raw in raw_values:
        if not raw or raw == "Any format":
            continue
        if raw in FORMAT_BY_ID:
            item_id = raw
        elif raw in FORMAT_BY_LABEL:
            item_id = str(FORMAT_BY_LABEL[raw]["id"])
        else:
            item_id = raw
        if item_id not in normalized:
            normalized.append(item_id)
    return normalized


def number_format_label(selected_format: str) -> str:
    selected_format = str(selected_format or "").strip()
    return FORMAT_LABEL_BY_ID.get(selected_format, selected_format)


def matches_number_format(number: str, selected_format: str) -> bool:
    number = str(number or "").strip()
    digits = "".join(ch for ch in number if ch.isdigit())
    selected_format = selected_format or "Any format"

    if selected_format == "Any format":
        return True

    meta = FORMAT_BY_ID.get(selected_format) or FORMAT_BY_LABEL.get(selected_format)
    if meta:
        required_length = int(meta.get("length") or 0) or None
        if required_length is not None and len(digits) != required_length:
            return False
        repeat = meta.get("repeat")
        if repeat:
            return any(digits.count(digit) >= int(repeat) for digit in set(digits))
        pattern = str(meta.get("pattern") or "")
        return _matches_pattern(digits, pattern) if pattern else True

    required_length = get_required_digit_length(selected_format)
    if required_length is not None and len(digits) != required_length:
        return False

    if selected_format == "Contains digit repeated 2 times":
        return any(digits.count(digit) >= 2 for digit in set(digits))
    if selected_format == "Contains digit repeated 3 times":
        return any(digits.count(digit) >= 3 for digit in set(digits))
    if selected_format == "Contains digit repeated 4 times":
        return any(digits.count(digit) >= 4 for digit in set(digits))

    pattern = get_format_pattern(selected_format)
    if not pattern:
        return True
    if len(digits) != len(pattern):
        if required_length is None or len(pattern) > len(digits):
            return False
        if pattern[0] == pattern[-1] and set(pattern[1:-1]) <= {"?"}:
            expanded_pattern = pattern[0] + ("?" * (len(digits) - 2)) + pattern[-1]
            return _matches_pattern(digits, expanded_pattern)
        return any(_matches_pattern(digits[index : index + len(pattern)], pattern) for index in range(0, len(digits) - len(pattern) + 1))

    return _matches_pattern(digits, pattern)


def match_number_formats(number: str, selected_formats: str | list[str] | tuple[str, ...] | None, fallback: str | None = None) -> dict[str, str | bool | list[str]]:
    formats = normalize_number_formats(selected_formats, fallback=fallback)
    selected_labels = [number_format_label(item) for item in formats]
    if not formats:
        return {
            "allowed": True,
            "matched": True,
            "matched_format": "",
            "matched_format_name": "",
            "selected_formats": [],
            "selected_format_labels": [],
            "skip_reason": "",
        }
    for item in formats:
        if matches_number_format(number, item):
            return {
                "allowed": True,
                "matched": True,
                "matched_format": item,
                "matched_format_name": number_format_label(item),
                "selected_formats": formats,
                "selected_format_labels": selected_labels,
                "skip_reason": "",
            }
    return {
        "allowed": False,
        "matched": False,
        "matched_format": "",
        "matched_format_name": "",
        "selected_formats": formats,
        "selected_format_labels": selected_labels,
        "skip_reason": "number format did not match selected formats",
    }


def _matches_pattern(digits: str, pattern: str) -> bool:
    if len(digits) != len(pattern):
        return False

    mapping: dict[str, str] = {}
    used_digits: dict[str, str] = {}
    for pattern_char, digit in zip(pattern, digits):
        if pattern_char == "?":
            continue
        if pattern_char in mapping:
            if mapping[pattern_char] != digit:
                return False
        else:
            if digit in used_digits and used_digits[digit] != pattern_char:
                return False
            mapping[pattern_char] = digit
            used_digits[digit] = pattern_char
    return True


def get_format_pattern(selected_format: str) -> str:
    if not selected_format or selected_format == "Any format":
        return ""
    meta = FORMAT_BY_ID.get(selected_format) or FORMAT_BY_LABEL.get(selected_format)
    if meta:
        return str(meta.get("pattern") or "")
    if selected_format.startswith("Contains digit repeated"):
        return ""
    return FORMAT_TO_URL_VALUE.get(selected_format, selected_format.split("(")[0].strip())


def get_required_digit_length(selected_format: str) -> int | None:
    meta = FORMAT_BY_ID.get(selected_format) or FORMAT_BY_LABEL.get(selected_format)
    if meta and meta.get("length"):
        return int(meta["length"])
    if "(5 Digits)" in selected_format:
        return 5
    if "(4 Digits)" in selected_format:
        return 4
    if "(3 Digits)" in selected_format:
        return 3
    if "(2 Digits)" in selected_format:
        return 2
    return None


def normalize_code(code: str | None) -> str:
    code = clean_text(code or "")
    return code.upper() if code else "?"


def build_xplate_url(
    city: str = "",
    code: str = "",
    contains: str = "",
    price_min: str = "",
    price_max: str = "",
    starts_with: str = "",
    ends_with: str = "",
    selected_format: str = "",
    page: int = 1,
) -> str:
    params = {
        "city": city_to_xplate_param(city),
        "code": code or "",
        "digits": "",
        "contains": contains or "",
        "price-max": price_max or "",
        "price-min": price_min or "",
        "starts-with": starts_with or "",
        "ends-with": ends_with or "",
        "format": get_format_pattern(selected_format) or "",
    }
    if page and page >= 1:
        params["page"] = page
    return BASE_URL + "?" + urlencode(params, doseq=False, safe="?+")


def build_url(
    city: str,
    number: str,
    search_mode: str = "contains",
    max_price: str = "",
    min_price: str = "",
    selected_format: str = "",
    page: int = 1,
) -> str:
    mode = normalize_search_mode(search_mode)
    if mode == "exact match":
        mode = "contains"
    contains = starts_with = ends_with = ""
    if mode == "starts with":
        starts_with = number
    elif mode == "ends with":
        ends_with = number
    else:
        contains = number
    return build_xplate_url(
        city=city,
        contains=contains,
        price_min=min_price,
        price_max=max_price,
        starts_with=starts_with,
        ends_with=ends_with,
        selected_format=selected_format,
        page=page,
    )


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(HEADERS)
    retries = Retry(
        total=1,
        connect=1,
        read=1,
        backoff_factor=0.5,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET",),
    )
    adapter = HTTPAdapter(max_retries=retries)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


def fetch_html(session: requests.Session, url: str, timeout: int = 10) -> str:
    response = session.get(url, timeout=timeout)
    response.raise_for_status()
    return response.text


def parse_soup(html: str) -> BeautifulSoup:
    return BeautifulSoup(html, "lxml")


def extract_listing_links(html: str) -> list[str]:
    soup = parse_soup(html)
    return extract_listing_links_from_seller_page(soup)


def extract_listing_links_from_seller_page(soup: BeautifulSoup) -> list[str]:
    links: list[str] = []
    seen: set[str] = set()

    for anchor in soup.find_all("a", href=True):
        href = anchor["href"].strip()
        absolute = urljoin(BASE_URL, href)
        parsed = urlparse(absolute)
        path = unquote(parsed.path)
        if "/en/numbers/license-plates/" not in path:
            continue
        if not re.search(r"/\d+-.+-plate-number-\d+$", path, re.IGNORECASE):
            continue
        clean_link = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
        if clean_link not in seen:
            seen.add(clean_link)
            links.append(clean_link)

    return links


def extract_listing_id_from_url(url: str) -> str:
    match = re.search(r"/license-plates/(\d+)(?:[-/]|$)", str(url or ""))
    return match.group(1) if match else ""


def _looks_like_listing_url(url: str) -> bool:
    path = unquote(urlparse(url).path)
    return "/en/numbers/license-plates/" in path and bool(re.search(r"/\d+-.+-plate-number-\d+", path, re.I))


def _listing_card_for_anchor(anchor):
    node = anchor
    best = anchor
    for _ in range(8):
        if not node or not getattr(node, "parent", None):
            break
        node = node.parent
        text = clean_text(node.get_text(" ", strip=True))
        links = [urljoin(BASE_URL, a.get("href", "")) for a in node.find_all("a", href=True)]
        listing_links = [link for link in links if _looks_like_listing_url(link)]
        has_signal = any(re.search(pattern, text, re.I) for pattern in (r"\bAED\b", r"\bcontact\b", r"\bfeatured\b", r"\bsold\b", r"\b\d+\s+(?:second|minute|hour|day)s?\s+ago\b", r"\btoday\b"))
        if len(listing_links) == 1 and (has_signal or len(text) > 40):
            best = node
            if has_signal:
                break
        if len(listing_links) > 1:
            break
    return best


def extract_listing_cards(soup: BeautifulSoup, source_url: str = "") -> list[dict[str, str]]:
    rows_by_id: dict[str, dict[str, str]] = {}
    rows_by_url: dict[str, dict[str, str]] = {}
    anchors = []
    for anchor in soup.find_all("a", href=True):
        absolute = urljoin(BASE_URL, anchor["href"].strip())
        if _looks_like_listing_url(absolute):
            anchors.append(anchor)

    for anchor in anchors:
        absolute = urljoin(BASE_URL, anchor["href"].strip())
        parsed = urlparse(absolute)
        link = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
        card = _listing_card_for_anchor(anchor)
        text = clean_text(card.get_text(" ", strip=True))
        row = row_from_listing_link(link, source_url=source_url)
        row.update(extract_card_metadata(text))
        row["listing_id"] = extract_listing_id_from_url(link)
        row["listing_link"] = link
        row["source_url"] = source_url
        seller_link = extract_seller_link(card)
        if seller_link:
            row["seller_link"] = seller_link
            row["seller_username"] = extract_seller_username(seller_link)
        listing_id = row.get("listing_id") or ""
        target = rows_by_id if listing_id else rows_by_url
        key = listing_id or link
        existing = target.get(key)
        if not existing:
            target[key] = row
        else:
            for field, value in row.items():
                if clean_text(str(value)) and str(existing.get(field, "")) in {"", "?", "Unknown", "Not available"}:
                    existing[field] = value
    return [*rows_by_id.values(), *rows_by_url.values()]


def extract_pagination_urls(soup: BeautifulSoup, current_url: str, max_pages: int = 3) -> list[str]:
    urls = [current_url]
    seen = {current_url}
    for anchor in soup.find_all("a", href=True):
        href = anchor["href"].strip()
        absolute = urljoin(current_url, href)
        parsed = urlparse(absolute)
        if "/en/numbers/users/" not in parsed.path or "/license-plates" not in parsed.path:
            continue
        query = parse_qs(parsed.query)
        page_values = query.get("page", [])
        if not page_values:
            continue
        try:
            page_number = int(page_values[0])
        except ValueError:
            continue
        if page_number < 1 or page_number > max_pages:
            continue
        clean_url = urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", urlencode({"page": page_number}), ""))
        if clean_url not in seen:
            seen.add(clean_url)
            urls.append(clean_url)
    return urls[:max_pages]


def has_next_page(soup: BeautifulSoup, current_page: int) -> bool:
    for anchor in soup.find_all("a", href=True):
        text = clean_text(anchor.get_text(" ", strip=True)).lower()
        href = anchor["href"]
        parsed = urlparse(urljoin(BASE_URL, href))
        query = parse_qs(parsed.query)
        page_values = query.get("page", [])
        if text in {"next", ">", "›", "»"}:
            return True
        if page_values:
            try:
                if int(page_values[0]) > current_page:
                    return True
            except ValueError:
                continue
    return False


def search_depth_to_max_pages(search_depth: str | None) -> int:
    if search_depth == "First page only":
        return 1
    if search_depth == "First 5 pages":
        return 5
    if search_depth == "First 10 pages":
        return 10
    return 100


def _slug_from_listing_link(link: str) -> str:
    return unquote(urlparse(link).path).rstrip("/").split("/")[-1]


def extract_plate_number(value: str) -> str:
    text = _slug_from_listing_link(value) if value.startswith("http") else value
    match = re.search(r"plate-number-(\d+)", text, re.IGNORECASE)
    return match.group(1) if match else ""


def extract_code(value: str) -> str:
    text = _slug_from_listing_link(value) if value.startswith("http") else value
    match = re.search(r"-code(?:-([a-zA-Z0-9]+))?-plate-number-", text, re.IGNORECASE)
    return normalize_code(match.group(1) if match else "")


def extract_city(value: str) -> str:
    text = _slug_from_listing_link(value) if value.startswith("http") else value
    match = re.match(r"\d+-(?P<city>.+)-code(?:-[a-zA-Z0-9]+)?-plate-number-\d+$", text, re.IGNORECASE)
    return match.group("city").replace("-", " ").title() if match else ""


def extract_price(text: str) -> str:
    text = clean_text(text)
    match = re.search(r"\bAED\s*([0-9]{1,3}(?:,[0-9]{3})*|[0-9]+)", text, re.IGNORECASE)
    if match:
        return "AED " + match.group(1)
    match = re.search(r"\b(Call For Price|Price hidden|Hidden price|Price on request)\b", text, re.IGNORECASE)
    if match:
        phrase = clean_text(match.group(1)).lower()
        return "Price hidden" if "hidden" in phrase or "request" in phrase or "call" in phrase else clean_text(match.group(1)).title()
    return ""


def price_to_number(price: str) -> float | None:
    match = re.search(r"([0-9][0-9,]*)", str(price or ""))
    if not match:
        return None
    return float(match.group(1).replace(",", ""))


def normalize_phone(text: str) -> str:
    text = clean_text(text)
    masked = re.search(r"\b0[1-9][0-9\s-]*[0-9]\s*x{2,4}\b", text, re.IGNORECASE)
    if masked:
        return clean_text(masked.group(0)).replace("-", "").replace(" ", "")
    patterns = [
        r"\+971[\s-]*\d[\d\s-]{7,12}",
        r"\b971[\s-]*\d[\d\s-]{7,12}\b",
        r"\b0[1-9][\d\s-]{3,12}x{2,4}\b",
        r"\b0[1-9][\d\s-]{6,10}\b",
    ]
    candidates: list[str] = []
    for pattern in patterns:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            phone = clean_text(match.group(0)).replace("-", "").replace(" ", "")
            if phone.startswith("971"):
                phone = "+" + phone
            candidates.append(phone)

    if not candidates:
        return "?"
    full_numbers = [phone for phone in candidates if "x" not in phone.lower()]
    return full_numbers[0] if full_numbers else candidates[0]


def extract_phone_from_text(text: str) -> str:
    text = clean_text(text)
    contact = re.search(r"\bContact\s*:?\s*([+0-9xX\s-]{7,28})", text, re.IGNORECASE)
    if contact:
        return normalize_phone(contact.group(1))
    show_contact = re.search(r"Show contact\s*([+0-9xX\s-]{7,24})", text, re.IGNORECASE)
    if show_contact:
        return normalize_phone(show_contact.group(1))
    profile_contact = re.search(r"contact me at\s*([+0-9\s-]{8,24})", text, re.IGNORECASE)
    if profile_contact:
        return normalize_phone(profile_contact.group(1))
    return normalize_phone(text)


def extract_seller_link(soup: BeautifulSoup) -> str:
    for anchor in soup.find_all("a", href=True):
        href = anchor["href"].strip()
        if "/en/numbers/users/" in href and "/license-plates" in href:
            return urljoin(BASE_URL, href)
    return ""


def extract_seller_username(seller_link: str) -> str:
    if not seller_link:
        return "?"
    match = re.search(r"/numbers/users/([^/]+)/license-plates", urlparse(seller_link).path)
    return unquote(match.group(1)) if match else "?"


def extract_seller_name(soup: BeautifulSoup) -> str:
    text = clean_text(soup.get_text(" ", strip=True))
    by_match = re.search(
        r"\bBy\s+(.+?)\s+(?:20\d{2}-\d{2}-\d{2}\s+\d{2}:\d{2}(?::\d{2})?)",
        text,
        re.IGNORECASE,
    )
    if by_match:
        return clean_text(by_match.group(1))

    seller_link = extract_seller_link(soup)
    if seller_link:
        for anchor in soup.find_all("a", href=True):
            if urljoin(BASE_URL, anchor["href"].strip()) != seller_link:
                continue
            seller_name = clean_text(anchor.get_text(" ", strip=True))
            if seller_name and "license plates" not in seller_name.lower():
                return seller_name
    return "Unknown"


def extract_uploaded_datetime(soup: BeautifulSoup) -> tuple[str, str]:
    text = clean_text(soup.get_text(" ", strip=True))
    match = re.search(r"\b(20\d{2}-\d{2}-\d{2})\s+(\d{2}:\d{2}(?::\d{2})?)\b", text)
    if match:
        return match.group(1), match.group(2)
    return "Not available", "Not available"


def extract_age_text(text: str) -> str:
    text = clean_text(text)
    match = re.search(r"\b(\d+\s+(?:second|seconds|minute|minutes|hour|hours|day|days|week|weeks|month|months|year|years)\s+ago)\b", text, re.I)
    if match:
        return clean_text(match.group(1))
    match = re.search(r"\b(today|yesterday)\b", text, re.I)
    return clean_text(match.group(1)) if match else "?"


def is_recent_age_text(age_text: str, window_minutes: int = 15) -> bool:
    text = clean_text(age_text).lower()
    match = re.search(r"\b(?:(\d+)\s+)?(second|seconds|minute|minutes)\s+ago\b", text, re.I)
    if not match:
        return False
    value = int(match.group(1) or 1)
    minutes = 0 if match.group(2).lower().startswith("second") else value
    return minutes <= max(int(window_minutes or 15), 1)


def extract_card_metadata(text: str) -> dict[str, str]:
    text = clean_text(text)
    featured = bool(re.search(r"\b(featured|promoted|sponsored)\b", text, re.I))
    sold = bool(re.search(r"\bsold\b", text, re.I))
    if sold:
        listing_type = "sold"
    elif featured:
        listing_type = "featured"
    elif extract_age_text(text) != "?":
        listing_type = "fresh_normal"
    else:
        listing_type = "old_static_promoted"
    views_match = re.search(r"\b(\d[\d,]*)\s+views?\b", text, re.I)
    return {
        "price": extract_price(text) or "?",
        "phone_number": extract_phone_from_text(text) or "?",
        "age_text": extract_age_text(text) or "?",
        "featured": featured,
        "sold": sold,
        "listing_type": listing_type,
        "views": views_match.group(1) if views_match else "?",
    }


def extract_phone_number_from_seller_profile(
    seller_link: str,
    session: requests.Session | None = None,
) -> tuple[str, str]:
    if not seller_link:
        return "Not available", "Unknown"
    session = session or make_session()
    try:
        html = fetch_html(session, seller_link)
    except requests.RequestException:
        return "Not available", "Unknown"

    soup = parse_soup(html)
    text = clean_text(soup.get_text(" ", strip=True))
    name = "Unknown"
    profile_match = re.search(r"License plates of\s+(.+?)\s+Home\b", text, re.IGNORECASE)
    if profile_match:
        name = clean_text(profile_match.group(1))
    else:
        hello_match = re.search(r"Hello,\s*I.m\s+(.+?)\s*,\s*the publisher", text, re.IGNORECASE)
        if hello_match:
            name = clean_text(hello_match.group(1))
    return extract_phone_from_text(text), name


def parse_listing_detail(
    listing_url: str,
    session: requests.Session | None = None,
    profile_cache: dict[str, tuple[str, str]] | None = None,
    timeout: int = 10,
) -> dict[str, str]:
    session = session or make_session()
    profile_cache = profile_cache if profile_cache is not None else {}
    html = fetch_html(session, listing_url, timeout=timeout)
    soup = parse_soup(html)
    text = clean_text(soup.get_text(" ", strip=True))
    uploaded_date, uploaded_time = extract_uploaded_datetime(soup)
    seller_link = extract_seller_link(soup)
    seller_username = extract_seller_username(seller_link)
    seller_name = extract_seller_name(soup)
    listing_phone = extract_phone_from_text(text)
    profile_phone, profile_name = "Not available", "Unknown"

    if seller_link:
        if seller_link not in profile_cache:
            profile_cache[seller_link] = extract_phone_number_from_seller_profile(seller_link, session)
        profile_phone, profile_name = profile_cache[seller_link]

    if seller_name == "Unknown" and profile_name != "Unknown":
        seller_name = profile_name
    phone_number = profile_phone if profile_phone != "Not available" else listing_phone

    row = {
        "city": extract_city(listing_url),
        "plate_number": extract_plate_number(listing_url),
        "code": extract_code(listing_url),
        "price": extract_price(text),
        "seller_name": seller_name,
        "seller_username": seller_username,
        "phone_number": phone_number,
        "uploaded_date": uploaded_date,
        "uploaded_time": uploaded_time,
        "age_text": extract_age_text(text),
        "featured": bool(re.search(r"\bfeatured\b", text, re.I)),
        "sold": bool(re.search(r"\bsold\b", text, re.I)),
        "listing_id": extract_listing_id_from_url(listing_url),
        "listing_link": listing_url,
        "seller_link": seller_link,
    }

    title_match = re.search(
        r"Plate number\s+(?P<city>[A-Za-z ]+?)\s+(?P<number>\d+)\s+code\s*(?P<code>[A-Za-z0-9]*)\s+for sale",
        text,
        re.IGNORECASE,
    )
    if title_match:
        row["city"] = title_match.group("city").title()
        row["plate_number"] = title_match.group("number")
        row["code"] = normalize_code(title_match.group("code"))

    return row


def row_from_listing_link(listing_link: str, source_url: str = "") -> dict[str, str]:
    return {
        "city": extract_city(listing_link),
        "plate_number": extract_plate_number(listing_link),
        "code": extract_code(listing_link),
        "price": "",
        "seller_name": "Unknown",
        "seller_username": "?",
        "phone_number": "?",
        "uploaded_date": "Not available",
        "uploaded_time": "Not available",
        "age_text": "Not available",
        "deal_rank": "Normal",
        "featured": False,
        "sold": False,
        "listing_type": "old_static_promoted",
        "views": "?",
        "listing_id": extract_listing_id_from_url(listing_link),
        "listing_link": listing_link,
        "seller_link": "",
        "source_url": source_url,
    }


def clean_result_row(row: dict[str, str]) -> dict[str, str]:
    cleaned = {column: row.get(column, "") for column in RESULT_COLUMNS}
    cleaned["code"] = normalize_code(cleaned.get("code"))
    if not clean_text(cleaned.get("price", "")):
        cleaned["price"] = "?"
    if not clean_text(cleaned.get("seller_name", "")) or cleaned.get("seller_name") == "Not available":
        cleaned["seller_name"] = "Unknown"
    if not clean_text(cleaned.get("seller_username", "")) or cleaned.get("seller_username") == "Not available":
        cleaned["seller_username"] = "?"
    if not clean_text(cleaned.get("phone_number", "")) or cleaned.get("phone_number") == "Not available":
        cleaned["phone_number"] = "?"
    cleaned["featured"] = bool(cleaned.get("featured"))
    cleaned["sold"] = bool(cleaned.get("sold"))
    return cleaned


def log_daily_listing_events(rows: list[dict[str, str]], debug_callback: Callable[[str], None] | None = None) -> None:
    """Record raw website listing events for daily Excel reports."""
    try:
        from . import plate_tracking

        for row in rows:
            plate_tracking.insert_listing_event(
                city=row.get("city", ""),
                plate_code=row.get("code", ""),
                plate_number=row.get("plate_number", ""),
                source="Website",
                price=row.get("price", ""),
                listing_url=row.get("listing_link", ""),
                raw_data_json=json.dumps(row, ensure_ascii=False, default=str),
            )
    except Exception as exc:
        debug_print(f"Failed to log listing events to tracking database: {exc}", debug_callback)


def matches_search_mode(plate_number: str, searched_number: str, search_mode: str) -> bool:
    mode = normalize_search_mode(search_mode)
    plate_number = str(plate_number or "")
    searched_number = str(searched_number or "")
    if mode == "exact match":
        return plate_number == searched_number
    if mode == "starts with":
        return plate_number.startswith(searched_number)
    if mode == "ends with":
        return plate_number.endswith(searched_number)
    return searched_number in plate_number


def filter_exact_matches(rows: list[dict[str, str]], searched_number: str) -> list[dict[str, str]]:
    return [row for row in rows if str(row.get("plate_number", "")) == str(searched_number)]


def apply_deal_rank(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    numeric_prices = [(index, price_to_number(row.get("price", ""))) for index, row in enumerate(rows)]
    numeric_prices = [(index, price) for index, price in numeric_prices if price is not None]
    for row in rows:
        row["deal_rank"] = "Normal"
    if numeric_prices:
        cheapest_index = min(numeric_prices, key=lambda item: item[1])[0]
        rows[cheapest_index]["deal_rank"] = "Cheapest"
    return rows


def sort_results(results: list[dict[str, str]], sort_mode: str = "Newest first") -> list[dict[str, str]]:
    def datetime_key(row: dict[str, str]):
        value = f"{row.get('uploaded_date', '')} {row.get('uploaded_time', '')}"
        try:
            return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return datetime.min

    mode = sort_mode or "Newest first"
    if mode == "Oldest first":
        return sorted(results, key=lambda row: (datetime_key(row) == datetime.min, datetime_key(row)))
    if mode == "Cheapest first":
        return sorted(results, key=lambda row: (price_to_number(row.get("price", "")) is None, price_to_number(row.get("price", "")) or 0))
    if mode in {"Highest price first", "Most expensive first"}:
        return sorted(results, key=lambda row: price_to_number(row.get("price", "")) or -1, reverse=True)
    if mode == "Seller name A-Z":
        return sorted(results, key=lambda row: row.get("seller_name", ""))
    if mode == "City A-Z":
        return sorted(results, key=lambda row: row.get("city", ""))
    if mode == "Code A-Z":
        return sorted(results, key=lambda row: row.get("code", ""))
    return sorted(results, key=lambda row: datetime_key(row), reverse=True)


def parse_seller_profile_cards(soup: BeautifulSoup, seller_link: str) -> list[dict[str, str]]:
    return extract_listing_cards(soup, seller_link)


def get_seller_plates(seller_link: str, max_pages: int = 3, timeout: int = 10) -> list[dict[str, str]]:
    if not seller_link:
        return []

    session = make_session()
    profile_cache: dict[str, tuple[str, str]] = {}
    listing_links: list[str] = []
    seen_links: set[str] = set()
    pages_to_fetch = [seller_link]
    seen_pages: set[str] = set()

    for page_url in pages_to_fetch[:max_pages]:
        if page_url in seen_pages:
            continue
        seen_pages.add(page_url)
        try:
            html = fetch_html(session, page_url, timeout=timeout)
        except requests.RequestException:
            continue

        soup = parse_soup(html)
        for extra_url in extract_pagination_urls(soup, page_url, max_pages=max_pages):
            if extra_url not in seen_pages and extra_url not in pages_to_fetch and len(pages_to_fetch) < max_pages:
                pages_to_fetch.append(extra_url)

        for row in parse_seller_profile_cards(soup, page_url):
            link = row.get("listing_link", "")
            if link and link not in seen_links:
                seen_links.add(link)
                listing_links.append(link)

    rows: list[dict[str, str]] = []
    for link in listing_links:
        try:
            detail = parse_listing_detail(link, session=session, profile_cache=profile_cache)
        except requests.RequestException:
            detail = row_from_listing_link(link, source_url=seller_link)
        rows.append(
            {
                "city": detail.get("city", ""),
                "plate_number": detail.get("plate_number", ""),
                "code": normalize_code(detail.get("code", "")),
                "price": detail.get("price", ""),
                "uploaded_date": detail.get("uploaded_date", "Not available"),
                "uploaded_time": detail.get("uploaded_time", "Not available"),
                "age_text": detail.get("age_text", "Not available"),
                "listing_link": detail.get("listing_link", link),
            }
        )

    return sort_results(rows, "Newest first")


def search_xplate(
    number: str,
    search_mode: str = "contains",
    code: str = "",
    contains: str = "",
    starts_with: str = "",
    ends_with: str = "",
    max_price: str = "",
    min_price: str = "",
    cities: list[str] | None = None,
    number_format: str = "Any format",
    number_formats: list[str] | None = None,
    search_depth: str = "All pages",
    sort_mode: str = "Newest first",
    delay_seconds: float = 0,
    debug_callback: Callable[[str], None] | None = None,
    progress_callback: Callable[[dict], None] | None = None,
    collect_details: bool = True,
    detail_timeout: int = 10,
    max_listings: int = 100,
    max_pages_override: int | None = None,
    auto_recent_window_minutes: int | None = None,
    auto_stop_after_old_pages: int = 2,
    **_unused,
) -> list[dict[str, str]]:
    number = clean_text(number)
    mode = normalize_search_mode(search_mode)
    if not number and mode == "exact match":
        mode = "contains"
    direct_contains = clean_text(contains)
    direct_starts_with = clean_text(starts_with)
    direct_ends_with = clean_text(ends_with)
    if not any([direct_contains, direct_starts_with, direct_ends_with]):
        if mode == "starts with":
            direct_starts_with = number
        elif mode == "ends with":
            direct_ends_with = number
        elif number:
            direct_contains = number

    selected_cities = [city for city in (cities or []) if clean_text(city).lower() not in {"", "all cities", "all"}]
    if not selected_cities:
        selected_cities = [""]
    session = make_session()
    rows: list[dict[str, str]] = []
    global_seen_page_links: set[str] = set()
    max_pages = int(max_pages_override or search_depth_to_max_pages(search_depth))
    selected_number_formats = normalize_number_formats(number_formats, fallback=number_format)
    url_number_format = selected_number_formats[0] if len(selected_number_formats) == 1 else "Any format"

    debug_print(f"Selected formats: {', '.join(number_format_label(item) for item in selected_number_formats) if selected_number_formats else 'Any format'}", debug_callback)
    debug_print(f"URL format value: {get_format_pattern(url_number_format) or '(none)'}", debug_callback)
    debug_print(f"Required digit length: {get_required_digit_length(url_number_format) or '(none)'}", debug_callback)

    total_city_count = max(len(selected_cities), 1)
    page_plan = max_pages if search_depth != "All pages" else None
    started_at = time.time()

    for city_index, city in enumerate(selected_cities, start=1):
        city_label = city.title() if city else "All cities"
        debug_print(f"Searching {city_label}", debug_callback)
        debug_print(f"Xplate city param: {city_to_xplate_param(city)}", debug_callback)
        city_page = 1
        recent_seen_for_city = False
        old_pages_after_recent = 0
        while city_page <= max_pages:
            url = build_xplate_url(
                city=city,
                code=code,
                contains=direct_contains,
                price_min=min_price,
                price_max=max_price,
                starts_with=direct_starts_with,
                ends_with=direct_ends_with,
                selected_format=url_number_format,
                page=city_page,
            )
            debug_print(f"Scraping URL: {url}", debug_callback)
            if progress_callback:
                pages_done = (city_index - 1) * (page_plan or 0) + city_page - 1
                pages_total = total_city_count * page_plan if page_plan else None
                progress_callback({
                    "current_city": city_label,
                    "current_page": city_page,
                    "pages_done": pages_done,
                    "pages_total": pages_total,
                    "results_so_far": len(rows),
                    "progress_percent": int((pages_done / pages_total) * 100) if pages_total else None,
                    "estimated_seconds_remaining": None,
                    "message": f"Searching {city_label}, page {city_page}...",
                })
            try:
                html = fetch_html(session, url)
            except requests.RequestException as exc:
                debug_print(f"Search page failed for {city}, page {city_page}: {exc}", debug_callback)
                break

            soup = parse_soup(html)
            card_rows = extract_listing_cards(soup, url)
            debug_print(f"Page {city_page} listing cards: {len(card_rows)}", debug_callback)
            if not card_rows and auto_recent_window_minutes is not None:
                break
            if not card_rows and max_pages_override is None:
                break
            if auto_recent_window_minutes is not None:
                page_has_recent = any(is_recent_age_text(row.get("age_text", ""), auto_recent_window_minutes) for row in card_rows)
                if page_has_recent:
                    recent_seen_for_city = True
                    old_pages_after_recent = 0
                elif recent_seen_for_city:
                    old_pages_after_recent += 1

            before_count = len(rows)
            new_links_on_page = 0
            for card_row in card_rows:
                link = card_row.get("listing_link", "")
                if link in global_seen_page_links:
                    continue
                global_seen_page_links.add(link)
                new_links_on_page += 1
                row = row_from_listing_link(link, source_url=url)
                row.update({key: value for key, value in card_row.items() if value not in ("", None)})
                if number and not matches_search_mode(row["plate_number"], number, mode):
                    continue
                if direct_contains and direct_contains not in str(row.get("plate_number", "")):
                    continue
                if direct_starts_with and not str(row.get("plate_number", "")).startswith(direct_starts_with):
                    continue
                if direct_ends_with and not str(row.get("plate_number", "")).endswith(direct_ends_with):
                    continue
                if code and normalize_code(row.get("code")) != normalize_code(code):
                    continue
                else:
                    rows.append(row)
            debug_print(f"Loaded {len(rows)} total raw results so far", debug_callback)
            if progress_callback:
                elapsed = max(time.time() - started_at, 0.1)
                pages_done = (city_index - 1) * (page_plan or city_page) + city_page
                pages_total = total_city_count * page_plan if page_plan else None
                remaining = None
                percent = None
                if pages_total:
                    average = elapsed / max(pages_done, 1)
                    remaining = max(int((pages_total - pages_done) * average), 0)
                    percent = min(int((pages_done / pages_total) * 100), 99)
                progress_callback({
                    "current_city": city_label,
                    "current_page": city_page,
                    "pages_done": pages_done,
                    "pages_total": pages_total,
                    "results_so_far": len(rows),
                    "progress_percent": percent,
                    "estimated_seconds_remaining": remaining,
                    "message": f"Loaded {len(rows)} results so far",
                })
            if new_links_on_page == 0 and max_pages_override is None:
                debug_print("No new listing links on this page; stopping pagination for this city", debug_callback)
                break

            if auto_recent_window_minutes is not None and not recent_seen_for_city and city_page >= 3:
                debug_print("No recent listing text found in the first automatic alert pages; stopping safely", debug_callback)
                break

            if auto_recent_window_minutes is not None and recent_seen_for_city and old_pages_after_recent >= max(auto_stop_after_old_pages, 1):
                debug_print("Recent listings already found and older pages started; stopping automatic alert scan", debug_callback)
                break

            if city_page >= max_pages:
                break
            if max_pages_override is None:
                if not has_next_page(soup, city_page) and len(rows) == before_count:
                    break
                if not has_next_page(soup, city_page) and len(card_rows) < 20:
                    break
            city_page += 1
    unique_rows: list[dict[str, str]] = []
    seen_links: set[str] = set()
    for row in rows:
        link = row["listing_link"]
        if link in seen_links:
            continue
        seen_links.add(link)
        unique_rows.append(row)

    if max_listings and len(unique_rows) > max_listings:
        debug_print(f"Limiting unique listings after full page scan to max listings per scan: {max_listings}", debug_callback)
        unique_rows = unique_rows[:max_listings]

    debug_print("Filtering results", debug_callback)
    if mode == "exact match":
        before_count = len(unique_rows)
        unique_rows = filter_exact_matches(unique_rows, number)
        debug_print(f"Exact-match filter removed {before_count - len(unique_rows)} rows", debug_callback)

    if selected_number_formats:
        before_count = len(unique_rows)
        unique_rows = [
            row for row in unique_rows
            if match_number_formats(row.get("plate_number", ""), selected_number_formats).get("matched")
        ]
        debug_print(f"Final after local filtering: {len(unique_rows)}", debug_callback)
        debug_print(f"Number-format filter removed {before_count - len(unique_rows)} rows", debug_callback)

    if not collect_details:
        debug_print("Detail enrichment skipped; using listing-card data only", debug_callback)
        final_rows = []
        for row in unique_rows:
            final_rows.append(clean_result_row(row))
        final_rows = apply_deal_rank(final_rows)
        final_rows = sort_results(final_rows, sort_mode)
        log_daily_listing_events(final_rows, debug_callback)
        debug_print(f"Done. Final results: {len(final_rows)}", debug_callback)
        return final_rows

    debug_print("Collecting seller details", debug_callback)
    profile_cache: dict[str, tuple[str, str]] = {}
    final_rows: list[dict[str, str]] = []
    for row in unique_rows:
        try:
            details = parse_listing_detail(row["listing_link"], session=session, profile_cache=profile_cache, timeout=detail_timeout)
            row.update({key: value for key, value in details.items() if value})
        except requests.RequestException as exc:
            debug_print(f"Detail page skipped: {row['listing_link']} ({exc})", debug_callback)

        final_rows.append(clean_result_row(row))
        if delay_seconds:
            time.sleep(delay_seconds)

    final_rows = apply_deal_rank(final_rows)
    final_rows = sort_results(final_rows, sort_mode)
    debug_print(f"Done. Final results: {len(final_rows)}", debug_callback)
    log_daily_listing_events(final_rows, debug_callback)
        
    return final_rows


def save_results(results: list[dict[str, str]], number: str):
    output_dir = Path("results")
    output_dir.mkdir(exist_ok=True)

    df = pd.DataFrame(results)
    df = pd.DataFrame(columns=RESULT_COLUMNS) if df.empty else df.reindex(columns=RESULT_COLUMNS)

    safe_number = re.sub(r"[^0-9A-Za-z_-]+", "_", number).strip("_") or "search"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    csv_path = output_dir / f"xplate_results_{safe_number}_{timestamp}.csv"
    xlsx_path = output_dir / f"xplate_results_{safe_number}_{timestamp}.xlsx"

    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    df.to_excel(xlsx_path, index=False)
    format_excel_file(xlsx_path)
    return df, csv_path, xlsx_path


def format_excel_file(path: Path) -> None:
    workbook = load_workbook(path)
    sheet = workbook.active
    sheet.freeze_panes = "A2"
    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 55)
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Search Xplate plate numbers across UAE cities.")
    parser.add_argument("number", help="Plate number to search, for example 2007")
    parser.add_argument(
        "--mode",
        choices=["contains", "starts", "starts with", "ends", "ends with", "exact", "exact match"],
        default="contains",
    )
    parser.add_argument("--max-price", default="")
    parser.add_argument("--min-price", default="")
    parser.add_argument("--sort", default="Newest first")
    args = parser.parse_args()

    results = search_xplate(
        args.number,
        search_mode=args.mode,
        max_price=args.max_price,
        min_price=args.min_price,
        number_format="Any format",
        sort_mode=args.sort,
    )
    df, csv_path, xlsx_path = save_results(results, args.number)

    print("\nDone.")
    print(f"Total results: {len(df)}")
    print(f"CSV: {csv_path}")
    print(f"Excel: {xlsx_path}")
    if not df.empty:
        print(df.to_string(index=False))


if __name__ == "__main__":
    main()
