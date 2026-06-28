from pathlib import Path

from openpyxl import load_workbook

from app import alerts
from app import plate_tracking
from app import storage


def _listing(**overrides):
    listing = {
        "city": "Dubai",
        "code": "A",
        "plate_number": "12345",
        "price": "AED 12,000",
        "listing_link": "https://example.test/listing/1",
        "source": "Website",
        "uploaded_date": "2026-06-25",
        "uploaded_time": "10:00",
    }
    listing.update(overrides)
    return listing


def test_rule_matches_are_deduplicated_and_workbook_is_grouped(tmp_path, monkeypatch):
    database_path = tmp_path / "plate_tracking.db"
    reports_path = tmp_path / "reports"
    monkeypatch.setattr(plate_tracking, "DB_PATH", database_path)
    monkeypatch.setattr(plate_tracking, "RULE_REPORTS_DIR", reports_path)
    monkeypatch.setattr(storage, "get_alerts", lambda: [{
        "id": "rule-1",
        "name": "Dubai VIP watch",
        "enabled": True,
        "cities": ["Dubai"],
        "code": "A",
        "price_max": "15000",
    }])
    plate_tracking.init_db()

    first = _listing()
    assert plate_tracking.insert_alert_rule_match(
        "rule-1", "Dubai VIP watch", first, "2026-06-25 10:01:00"
    )
    assert not plate_tracking.insert_alert_rule_match(
        "rule-1", "Dubai VIP watch", first, "2026-06-25 10:02:00"
    )
    assert plate_tracking.insert_alert_rule_match(
        "rule-1",
        "Dubai VIP watch",
        _listing(price="AED 11,500"),
        "2026-06-25 10:03:00",
    )
    assert plate_tracking.insert_alert_rule_match(
        "rule-1",
        "Dubai VIP watch",
        _listing(listing_link="https://example.test/listing/2", uploaded_time="11:00"),
        "2026-06-25 11:01:00",
    )
    assert plate_tracking.insert_alert_rule_match(
        "rule-1",
        "Dubai VIP watch",
        _listing(
            city="Sharjah",
            code="3",
            plate_number="7777",
            price="",
            listing_link="https://example.test/listing/3",
            source="OCR Instagram",
        ),
        "2026-06-25 12:01:00",
    )

    report_path = plate_tracking.generate_daily_rule_excel_report("rule-1", "2026-06-25")
    assert report_path == Path(reports_path / "rule-1" / "XPLATE REPORT 2026-06-25.xlsx").resolve()
    workbook = load_workbook(report_path)
    assert workbook.sheetnames == ["Summary", "Dubai", "Sharjah"]

    headers = [cell.value for cell in workbook["Dubai"][1]]
    assert headers == [
        "Full Plate",
        "Digits",
        "Source",
        "Times Uploaded Today",
        "Price",
        "All Prices Seen",
        "Listing Links",
        "Notes",
    ]
    dubai_row = [cell.value for cell in workbook["Dubai"][2]]
    assert dubai_row[0] == "Dubai A 12345"
    assert dubai_row[3] == 3
    assert dubai_row[4] == 12000
    assert dubai_row[5] == "AED 12,000, AED 11,500, AED 12,000"
    assert workbook["Dubai"].freeze_panes == "A2"
    assert workbook["Dubai"].auto_filter.ref == "A1:H2"

    summary_values = {
        workbook["Summary"].cell(row=row, column=1).value: workbook["Summary"].cell(row=row, column=2).value
        for row in range(1, workbook["Summary"].max_row + 1)
    }
    assert summary_values["Report type"] == "Daily Saved Rule Report"
    assert summary_values["Rule ID"] == "rule-1"
    assert summary_values["Total unique plates"] == 2
    assert summary_values["Total listing events"] == 4
    assert summary_values["Total repeated plates"] == 1


def test_no_data_report_has_summary_only(tmp_path, monkeypatch):
    monkeypatch.setattr(plate_tracking, "DB_PATH", tmp_path / "plate_tracking.db")
    monkeypatch.setattr(plate_tracking, "RULE_REPORTS_DIR", tmp_path / "reports")
    monkeypatch.setattr(storage, "get_alerts", lambda: [{
        "id": "rule-empty",
        "name": "Empty rule",
        "enabled": True,
    }])
    plate_tracking.init_db()

    report_path = plate_tracking.generate_daily_rule_excel_report("rule-empty", "2026-06-25")
    workbook = load_workbook(report_path)
    assert workbook.sheetnames == ["Summary"]
    assert workbook["Summary"]["A4"].value == "No data found for this saved rule on the selected date."


def test_general_daily_report_uses_selected_date_and_allowed_columns(tmp_path, monkeypatch):
    monkeypatch.setattr(plate_tracking, "DB_PATH", tmp_path / "plate_tracking.db")
    monkeypatch.setattr(plate_tracking, "REPORTS_DIR", tmp_path / "daily")
    plate_tracking.init_db()
    plate_tracking.insert_listing_event(
        city="Dubai",
        plate_code="A",
        plate_number="12345",
        source="Website",
        price="AED 12,000",
        listing_url="https://example.test/1",
        seen_at="2026-06-28 08:00:00",
    )
    plate_tracking.insert_listing_event(
        city="Sharjah",
        plate_code="3",
        plate_number="7777",
        source="Instagram",
        price="",
        listing_url="https://example.test/2",
        seen_at="2026-06-27 23:59:59",
    )

    report_path = Path(plate_tracking.generate_daily_excel_report("2026-06-28"))
    workbook = load_workbook(report_path)

    assert report_path.name == "XPLATE REPORT 2026-06-28.xlsx"
    assert workbook.sheetnames == ["Summary", "Dubai"]
    assert [cell.value for cell in workbook["Dubai"][1]] == [
        "Full Plate",
        "Digits",
        "Source",
        "Times Uploaded Today",
        "Price",
        "All Prices Seen",
        "Listing Links",
        "Notes",
    ]
    assert workbook["Dubai"]["A2"].value == "Dubai A 12345"


def test_general_no_data_report_has_clear_summary(tmp_path, monkeypatch):
    monkeypatch.setattr(plate_tracking, "DB_PATH", tmp_path / "plate_tracking.db")
    monkeypatch.setattr(plate_tracking, "REPORTS_DIR", tmp_path / "daily")
    plate_tracking.init_db()

    report_path = plate_tracking.generate_daily_excel_report("2026-06-28")
    workbook = load_workbook(report_path)

    assert workbook.sheetnames == ["Summary"]
    assert workbook["Summary"]["A4"].value == "No data found for this date"


def test_pin_telegram_message_uses_same_chat_and_silent_notification(monkeypatch):
    captured = {}

    class Response:
        ok = True
        status_code = 200
        text = ""

        @staticmethod
        def json():
            return {"ok": True, "result": True}

    def fake_post(url, json, timeout):
        captured.update({"url": url, "payload": json, "timeout": timeout})
        return Response()

    monkeypatch.setattr(alerts.requests, "post", fake_post)
    assert alerts.pin_telegram_message("@eliteplates", 8511, "test-token")
    assert captured["url"].endswith("/bottest-token/pinChatMessage")
    assert captured["payload"] == {
        "chat_id": "@eliteplates",
        "message_id": 8511,
        "disable_notification": True,
    }
    assert captured["timeout"] == 30


def test_telegram_document_uses_exact_report_filename(tmp_path, monkeypatch):
    report_path = tmp_path / "XPLATE REPORT 2026-06-25.xlsx"
    report_path.write_bytes(b"test workbook")
    captured = {}

    class Response:
        @staticmethod
        def raise_for_status():
            return None

        @staticmethod
        def json():
            return {"ok": True, "result": {"message_id": 8511}}

    def fake_post(url, data, files, timeout):
        captured.update({
            "url": url,
            "data": data,
            "filename": files["document"][0],
            "timeout": timeout,
        })
        return Response()

    monkeypatch.setattr(alerts.requests, "post", fake_post)
    response = alerts.send_telegram_document(
        "test-token",
        "@eliteplates",
        str(report_path),
        "Daily report",
    )

    assert response["ok"] is True
    assert captured["filename"] == "XPLATE REPORT 2026-06-25.xlsx"


def test_daily_report_pin_failure_is_non_fatal(tmp_path, monkeypatch, caplog):
    rule = {"id": "rule-1", "name": "Dubai VIP watch", "enabled": True}
    monkeypatch.setenv("TELEGRAM_PIN_DAILY_REPORTS", "true")
    monkeypatch.setattr(alerts, "get_alerts", lambda: [rule])
    monkeypatch.setattr(alerts, "_resolve_telegram_credentials", lambda alert: ("test-token", "@eliteplates"))
    monkeypatch.setattr(
        alerts.plate_tracking,
        "generate_daily_rule_excel_report",
        lambda *args, **kwargs: tmp_path / "report.xlsx",
    )
    monkeypatch.setattr(
        alerts.plate_tracking,
        "aggregate_daily_rule_report",
        lambda *args, **kwargs: {
            "summary": {
                "total_unique_plates": 2,
                "total_listing_events": 3,
                "total_repeated_plates": 1,
            }
        },
    )
    monkeypatch.setattr(
        alerts,
        "send_telegram_document",
        lambda *args, **kwargs: {
            "ok": True,
            "result": {
                "message_id": 8511,
                "document": {"file_name": "XPLATE REPORT 2026-06-27.xlsx"},
            },
        },
    )
    monkeypatch.setattr(alerts, "pin_telegram_message", lambda *args, **kwargs: False)
    monkeypatch.setattr(alerts, "_daily_report_log", lambda *args, **kwargs: None)

    result = alerts.send_daily_rule_report_to_telegram("rule-1", "2026-06-27")

    assert result["ok"] is True
    assert result["telegram_message_id"] == 8511
    assert result["telegram_document_name"] == "XPLATE REPORT 2026-06-27.xlsx"
    assert result["telegram_pinning_enabled"] is True
    assert result["telegram_pin_attempted"] is True
    assert result["telegram_pinned"] is False
    assert "Telegram report sent, but pinning failed" in caplog.text


def test_daily_report_pinning_defaults_to_disabled(monkeypatch):
    monkeypatch.delenv("TELEGRAM_PIN_DAILY_REPORTS", raising=False)
    assert alerts.env_bool("TELEGRAM_PIN_DAILY_REPORTS", False) is False
