import argparse
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Callable
from urllib.parse import parse_qs, quote_plus, unquote, urlencode, urljoin, urlparse, urlunparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


BASE_URL = "https://xplate.com/en/numbers/license-plates"

CITIES = [
    "abu dhabi",
    "dubai",
    "sharjah",
    "ajman",
    "umm al quwain",
    "ras al khaimah",
    "fujairah",
]

RESULT_COLUMNS = [
    "city",
    "plate_number",
    "code",
    "price",
    "seller_name",
    "seller_username",
    "phone_number",
    "uploaded_date",
    "uploaded_time",
    "age_text",
    "deal_rank",
    "listing_link",
    "seller_link",
    "source_url",
]

NUMBER_FORMAT_OPTIONS = [
    "Any format",
    "Contains digit repeated 2 times",
    "Contains digit repeated 3 times",
    "Contains digit repeated 4 times",
    "x??x (5 Digits)",
    "xyzyx (5 Digits)",
    "xyxxx (5 Digits)",
    "?xxx? (5 Digits)",
    "xyxyx (5 Digits)",
    "xyyyx (5 Digits)",
    "??xxx (5 Digits)",
    "xxx?? (5 Digits)",
    "xyyy (5 Digits)",
    "xyxy (5 Digits)",
    "xxx (5 Digits)",
    "xxxy (5 Digits)",
    "xyxx (5 Digits)",
    "xxyx (5 Digits)",
    "xxxxx (5 Digits)",
    "x??x (4 Digits)",
    "xyyx (4 Digits)",
    "xyxy (4 Digits)",
    "xxyy (4 Digits)",
    "xxxy (4 Digits)",
    "xyxx (4 Digits)",
    "xxxx (4 Digits)",
    "x?x (3 Digits)",
    "xyx (3 Digits)",
    "xxy (3 Digits)",
    "xyy (3 Digits)",
    "xxx (3 Digits)",
    "xx (2 Digits)",
    "xy (2 Digits)",
]

FORMAT_TO_URL_VALUE = {
    "Any format": "",
    "Contains digit repeated 2 times": "",
    "Contains digit repeated 3 times": "",
    "Contains digit repeated 4 times": "",
    "x??x (5 Digits)": "x??x",
    "xyzyx (5 Digits)": "xyzyx",
    "xyxxx (5 Digits)": "xyxxx",
    "?xxx? (5 Digits)": "?xxx?",
    "xyxyx (5 Digits)": "xyxyx",
    "xyyyx (5 Digits)": "xyyyx",
    "??xxx (5 Digits)": "??xxx",
    "xxx?? (5 Digits)": "xxx??",
    "xyyy (5 Digits)": "xyyy",
    "xyxy (5 Digits)": "xyxy",
    "xxx (5 Digits)": "xxx",
    "xxxy (5 Digits)": "xxxy",
    "xyxx (5 Digits)": "xyxx",
    "xxyx (5 Digits)": "xxyx",
    "xxxxx (5 Digits)": "xxxxx",
    "x??x (4 Digits)": "x??x",
    "xyyx (4 Digits)": "xyyx",
    "xyxy (4 Digits)": "xyxy",
    "xxyy (4 Digits)": "xxyy",
    "xxxy (4 Digits)": "xxxy",
    "xyxx (4 Digits)": "xyxx",
    "xxxx (4 Digits)": "xxxx",
    "x?x (3 Digits)": "x?x",
    "xyx (3 Digits)": "xyx",
    "xxy (3 Digits)": "xxy",
    "xyy (3 Digits)": "xyy",
    "xxx (3 Digits)": "xxx",
    "xx (2 Digits)": "xx",
    "xy (2 Digits)": "xy",
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}


def debug_print(message: str, debug_callback: Callable[[str], None] | None = None) -> None:
    print(message)
    if debug_callback:
        debug_callback(message)


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def normalize_search_mode(search_mode: str) -> str:
    mode = (search_mode or "contains").strip().lower().replace("_", " ")
    if mode in {"starts", "start", "starts with", "starts-with"}:
        return "starts with"
    if mode in {"ends", "end", "ends with", "ends-with"}:
        return "ends with"
    if mode in {"exact", "exact match", "exact-match"}:
        return "exact match"
    return "contains"


def matches_number_format(number: str, selected_format: str) -> bool:
    number = str(number or "").strip()
    digits = "".join(ch for ch in number if ch.isdigit())
    selected_format = selected_format or "Any format"

    if selected_format == "Any format":
        return True

    required_length = get_required_digit_length(selected_format)
    if required_length is not None and len(digits) != required_length:
        return False

    if selected_format == "Contains digit repeated 2 times":
        return any(digits.count(digit) >= 2 for digit in set(digits))
    if selected_format == "Contains digit repeated 3 times":
        return any(digits.count(digit) >= 3 for digit in set(digits))
    if selected_format == "Contains digit repeated 4 times":
        return any(digits.count(digit) >= 4 for digit in set(digits))

    pattern = get_format_pattern(selected_format)
    if not pattern:
        return True
    if len(digits) != len(pattern):
        if required_length is None or len(pattern) > len(digits):
            return False
        if pattern[0] == pattern[-1] and set(pattern[1:-1]) <= {"?"}:
            expanded_pattern = pattern[0] + ("?" * (len(digits) - 2)) + pattern[-1]
            return _matches_pattern(digits, expanded_pattern)
        return any(_matches_pattern(digits[index : index + len(pattern)], pattern) for index in range(0, len(digits) - len(pattern) + 1))

    return _matches_pattern(digits, pattern)


def _matches_pattern(digits: str, pattern: str) -> bool:
    if len(digits) != len(pattern):
        return False

    mapping: dict[str, str] = {}
    used_digits: dict[str, str] = {}
    for pattern_char, digit in zip(pattern, digits):
        if pattern_char == "?":
            continue
        if pattern_char in mapping:
            if mapping[pattern_char] != digit:
                return False
        else:
            if digit in used_digits and used_digits[digit] != pattern_char:
                return False
            mapping[pattern_char] = digit
            used_digits[digit] = pattern_char
    return True


def get_format_pattern(selected_format: str) -> str:
    if not selected_format or selected_format == "Any format":
        return ""
    if selected_format.startswith("Contains digit repeated"):
        return ""
    return FORMAT_TO_URL_VALUE.get(selected_format, selected_format.split("(")[0].strip())


def get_required_digit_length(selected_format: str) -> int | None:
    if "(5 Digits)" in selected_format:
        return 5
    if "(4 Digits)" in selected_format:
        return 4
    if "(3 Digits)" in selected_format:
        return 3
    if "(2 Digits)" in selected_format:
        return 2
    return None


def normalize_code(code: str | None) -> str:
    code = clean_text(code or "")
    return code.upper() if code else "?"


def build_xplate_url(
    city: str = "",
    code: str = "",
    contains: str = "",
    price_min: str = "",
    price_max: str = "",
    starts_with: str = "",
    ends_with: str = "",
    selected_format: str = "",
    page: int = 1,
) -> str:
    params = {
        "city": city or "",
        "code": code or "",
        "digits": "",
        "contains": contains or "",
        "price-max": price_max or "",
        "price-min": price_min or "",
        "starts-with": starts_with or "",
        "ends-with": ends_with or "",
        "format": get_format_pattern(selected_format) or "",
    }
    if page and page > 1:
        params["page"] = page
    return BASE_URL + "?" + urlencode(params, doseq=False, safe="?")


def build_url(
    city: str,
    number: str,
    search_mode: str = "contains",
    max_price: str = "",
    min_price: str = "",
    selected_format: str = "",
    page: int = 1,
) -> str:
    mode = normalize_search_mode(search_mode)
    if mode == "exact match":
        mode = "contains"
    contains = starts_with = ends_with = ""
    if mode == "starts with":
        starts_with = number
    elif mode == "ends with":
        ends_with = number
    else:
        contains = number
    return build_xplate_url(
        city=city,
        contains=contains,
        price_min=min_price,
        price_max=max_price,
        starts_with=starts_with,
        ends_with=ends_with,
        selected_format=selected_format,
        page=page,
    )


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(HEADERS)
    retries = Retry(
        total=1,
        connect=1,
        read=1,
        backoff_factor=0.5,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET",),
    )
    adapter = HTTPAdapter(max_retries=retries)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


def fetch_html(session: requests.Session, url: str, timeout: int = 10) -> str:
    response = session.get(url, timeout=timeout)
    response.raise_for_status()
    return response.text


def parse_soup(html: str) -> BeautifulSoup:
    return BeautifulSoup(html, "lxml")


def extract_listing_links(html: str) -> list[str]:
    soup = parse_soup(html)
    return extract_listing_links_from_seller_page(soup)


def extract_listing_links_from_seller_page(soup: BeautifulSoup) -> list[str]:
    links: list[str] = []
    seen: set[str] = set()

    for anchor in soup.find_all("a", href=True):
        href = anchor["href"].strip()
        absolute = urljoin(BASE_URL, href)
        parsed = urlparse(absolute)
        path = unquote(parsed.path)
        if "/en/numbers/license-plates/" not in path:
            continue
        if not re.search(r"/\d+-.+-plate-number-\d+$", path, re.IGNORECASE):
            continue
        clean_link = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
        if clean_link not in seen:
            seen.add(clean_link)
            links.append(clean_link)

    return links


def extract_pagination_urls(soup: BeautifulSoup, current_url: str, max_pages: int = 3) -> list[str]:
    urls = [current_url]
    seen = {current_url}
    for anchor in soup.find_all("a", href=True):
        href = anchor["href"].strip()
        absolute = urljoin(current_url, href)
        parsed = urlparse(absolute)
        if "/en/numbers/users/" not in parsed.path or "/license-plates" not in parsed.path:
            continue
        query = parse_qs(parsed.query)
        page_values = query.get("page", [])
        if not page_values:
            continue
        try:
            page_number = int(page_values[0])
        except ValueError:
            continue
        if page_number < 1 or page_number > max_pages:
            continue
        clean_url = urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", urlencode({"page": page_number}), ""))
        if clean_url not in seen:
            seen.add(clean_url)
            urls.append(clean_url)
    return urls[:max_pages]


def has_next_page(soup: BeautifulSoup, current_page: int) -> bool:
    for anchor in soup.find_all("a", href=True):
        text = clean_text(anchor.get_text(" ", strip=True)).lower()
        href = anchor["href"]
        parsed = urlparse(urljoin(BASE_URL, href))
        query = parse_qs(parsed.query)
        page_values = query.get("page", [])
        if text in {"next", ">", "›", "»"}:
            return True
        if page_values:
            try:
                if int(page_values[0]) > current_page:
                    return True
            except ValueError:
                continue
    return False


def search_depth_to_max_pages(search_depth: str | None) -> int:
    if search_depth == "First page only":
        return 1
    if search_depth == "First 5 pages":
        return 5
    if search_depth == "First 10 pages":
        return 10
    return 100


def _slug_from_listing_link(link: str) -> str:
    return unquote(urlparse(link).path).rstrip("/").split("/")[-1]


def extract_plate_number(value: str) -> str:
    text = _slug_from_listing_link(value) if value.startswith("http") else value
    match = re.search(r"plate-number-(\d+)", text, re.IGNORECASE)
    return match.group(1) if match else ""


def extract_code(value: str) -> str:
    text = _slug_from_listing_link(value) if value.startswith("http") else value
    match = re.search(r"-code(?:-([a-zA-Z0-9]+))?-plate-number-", text, re.IGNORECASE)
    return normalize_code(match.group(1) if match else "")


def extract_city(value: str) -> str:
    text = _slug_from_listing_link(value) if value.startswith("http") else value
    match = re.match(r"\d+-(?P<city>.+)-code(?:-[a-zA-Z0-9]+)?-plate-number-\d+$", text, re.IGNORECASE)
    return match.group("city").replace("-", " ").title() if match else ""


def extract_price(text: str) -> str:
    text = clean_text(text)
    match = re.search(r"\bAED\s*([0-9]{1,3}(?:,[0-9]{3})*|[0-9]+)", text, re.IGNORECASE)
    if match:
        return "AED " + match.group(1)
    match = re.search(r"\b(Call For Price|Price hidden)\b", text, re.IGNORECASE)
    return clean_text(match.group(1)).title() if match else ""


def price_to_number(price: str) -> float | None:
    match = re.search(r"([0-9][0-9,]*)", str(price or ""))
    if not match:
        return None
    return float(match.group(1).replace(",", ""))


def normalize_phone(text: str) -> str:
    text = clean_text(text)
    patterns = [
        r"\+971[\s-]*\d[\d\s-]{7,12}",
        r"\b971[\s-]*\d[\d\s-]{7,12}\b",
        r"\b0[1-9]\d{7,8}\b",
        r"\b0[1-9]\d{4,8}x{2,3}\b",
    ]
    candidates: list[str] = []
    for pattern in patterns:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            phone = clean_text(match.group(0)).replace("-", "").replace(" ", "")
            if phone.startswith("971"):
                phone = "+" + phone
            candidates.append(phone)

    if not candidates:
        return "Not available"
    full_numbers = [phone for phone in candidates if "x" not in phone.lower()]
    return full_numbers[0] if full_numbers else candidates[0]


def extract_phone_from_text(text: str) -> str:
    text = clean_text(text)
    show_contact = re.search(r"Show contact\s*([+0-9xX\s-]{7,24})", text, re.IGNORECASE)
    if show_contact:
        return normalize_phone(show_contact.group(1))
    profile_contact = re.search(r"contact me at\s*([+0-9\s-]{8,24})", text, re.IGNORECASE)
    if profile_contact:
        return normalize_phone(profile_contact.group(1))
    return normalize_phone(text)


def extract_seller_link(soup: BeautifulSoup) -> str:
    for anchor in soup.find_all("a", href=True):
        href = anchor["href"].strip()
        if "/en/numbers/users/" in href and "/license-plates" in href:
            return urljoin(BASE_URL, href)
    return ""


def extract_seller_username(seller_link: str) -> str:
    if not seller_link:
        return "Not available"
    match = re.search(r"/numbers/users/([^/]+)/license-plates", urlparse(seller_link).path)
    return unquote(match.group(1)) if match else "Not available"


def extract_seller_name(soup: BeautifulSoup) -> str:
    text = clean_text(soup.get_text(" ", strip=True))
    by_match = re.search(
        r"\bBy\s+(.+?)\s+(?:20\d{2}-\d{2}-\d{2}\s+\d{2}:\d{2}(?::\d{2})?)",
        text,
        re.IGNORECASE,
    )
    if by_match:
        return clean_text(by_match.group(1))

    seller_link = extract_seller_link(soup)
    if seller_link:
        for anchor in soup.find_all("a", href=True):
            if urljoin(BASE_URL, anchor["href"].strip()) != seller_link:
                continue
            seller_name = clean_text(anchor.get_text(" ", strip=True))
            if seller_name and "license plates" not in seller_name.lower():
                return seller_name
    return "Unknown"


def extract_uploaded_datetime(soup: BeautifulSoup) -> tuple[str, str]:
    text = clean_text(soup.get_text(" ", strip=True))
    match = re.search(r"\b(20\d{2}-\d{2}-\d{2})\s+(\d{2}:\d{2}(?::\d{2})?)\b", text)
    if match:
        return match.group(1), match.group(2)
    return "Not available", "Not available"


def extract_age_text(text: str) -> str:
    text = clean_text(text)
    match = re.search(r"\b(\d+\s+(?:minute|minutes|hour|hours|day|days|month|months|year|years)\s+ago)\b", text, re.I)
    return clean_text(match.group(1)) if match else "Not available"


def extract_phone_number_from_seller_profile(
    seller_link: str,
    session: requests.Session | None = None,
) -> tuple[str, str]:
    if not seller_link:
        return "Not available", "Unknown"
    session = session or make_session()
    try:
        html = fetch_html(session, seller_link)
    except requests.RequestException:
        return "Not available", "Unknown"

    soup = parse_soup(html)
    text = clean_text(soup.get_text(" ", strip=True))
    name = "Unknown"
    profile_match = re.search(r"License plates of\s+(.+?)\s+Home\b", text, re.IGNORECASE)
    if profile_match:
        name = clean_text(profile_match.group(1))
    else:
        hello_match = re.search(r"Hello,\s*I.m\s+(.+?)\s*,\s*the publisher", text, re.IGNORECASE)
        if hello_match:
            name = clean_text(hello_match.group(1))
    return extract_phone_from_text(text), name


def parse_listing_detail(
    listing_url: str,
    session: requests.Session | None = None,
    profile_cache: dict[str, tuple[str, str]] | None = None,
) -> dict[str, str]:
    session = session or make_session()
    profile_cache = profile_cache if profile_cache is not None else {}
    html = fetch_html(session, listing_url)
    soup = parse_soup(html)
    text = clean_text(soup.get_text(" ", strip=True))
    uploaded_date, uploaded_time = extract_uploaded_datetime(soup)
    seller_link = extract_seller_link(soup)
    seller_username = extract_seller_username(seller_link)
    seller_name = extract_seller_name(soup)
    listing_phone = extract_phone_from_text(text)
    profile_phone, profile_name = "Not available", "Unknown"

    if seller_link:
        if seller_link not in profile_cache:
            profile_cache[seller_link] = extract_phone_number_from_seller_profile(seller_link, session)
        profile_phone, profile_name = profile_cache[seller_link]

    if seller_name == "Unknown" and profile_name != "Unknown":
        seller_name = profile_name
    phone_number = profile_phone if profile_phone != "Not available" else listing_phone

    row = {
        "city": extract_city(listing_url),
        "plate_number": extract_plate_number(listing_url),
        "code": extract_code(listing_url),
        "price": extract_price(text),
        "seller_name": seller_name,
        "seller_username": seller_username,
        "phone_number": phone_number,
        "uploaded_date": uploaded_date,
        "uploaded_time": uploaded_time,
        "age_text": extract_age_text(text),
        "listing_link": listing_url,
        "seller_link": seller_link,
    }

    title_match = re.search(
        r"Plate number\s+(?P<city>[A-Za-z ]+?)\s+(?P<number>\d+)\s+code\s*(?P<code>[A-Za-z0-9]*)\s+for sale",
        text,
        re.IGNORECASE,
    )
    if title_match:
        row["city"] = title_match.group("city").title()
        row["plate_number"] = title_match.group("number")
        row["code"] = normalize_code(title_match.group("code"))

    return row


def row_from_listing_link(listing_link: str, source_url: str = "") -> dict[str, str]:
    return {
        "city": extract_city(listing_link),
        "plate_number": extract_plate_number(listing_link),
        "code": extract_code(listing_link),
        "price": "",
        "seller_name": "Unknown",
        "seller_username": "?",
        "phone_number": "?",
        "uploaded_date": "Not available",
        "uploaded_time": "Not available",
        "age_text": "Not available",
        "deal_rank": "Normal",
        "listing_link": listing_link,
        "seller_link": "",
        "source_url": source_url,
    }


def clean_result_row(row: dict[str, str]) -> dict[str, str]:
    cleaned = {column: row.get(column, "") for column in RESULT_COLUMNS}
    cleaned["code"] = normalize_code(cleaned.get("code"))
    if not clean_text(cleaned.get("price", "")):
        cleaned["price"] = "?"
    if not clean_text(cleaned.get("seller_name", "")) or cleaned.get("seller_name") == "Not available":
        cleaned["seller_name"] = "Unknown"
    if not clean_text(cleaned.get("seller_username", "")) or cleaned.get("seller_username") == "Not available":
        cleaned["seller_username"] = "?"
    if not clean_text(cleaned.get("phone_number", "")) or cleaned.get("phone_number") == "Not available":
        cleaned["phone_number"] = "?"
    return cleaned


def matches_search_mode(plate_number: str, searched_number: str, search_mode: str) -> bool:
    mode = normalize_search_mode(search_mode)
    plate_number = str(plate_number or "")
    searched_number = str(searched_number or "")
    if mode == "exact match":
        return plate_number == searched_number
    if mode == "starts with":
        return plate_number.startswith(searched_number)
    if mode == "ends with":
        return plate_number.endswith(searched_number)
    return searched_number in plate_number


def filter_exact_matches(rows: list[dict[str, str]], searched_number: str) -> list[dict[str, str]]:
    return [row for row in rows if str(row.get("plate_number", "")) == str(searched_number)]


def apply_deal_rank(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    numeric_prices = [(index, price_to_number(row.get("price", ""))) for index, row in enumerate(rows)]
    numeric_prices = [(index, price) for index, price in numeric_prices if price is not None]
    for row in rows:
        row["deal_rank"] = "Normal"
    if numeric_prices:
        cheapest_index = min(numeric_prices, key=lambda item: item[1])[0]
        rows[cheapest_index]["deal_rank"] = "Cheapest"
    return rows


def sort_results(results: list[dict[str, str]], sort_mode: str = "Newest first") -> list[dict[str, str]]:
    def datetime_key(row: dict[str, str]):
        value = f"{row.get('uploaded_date', '')} {row.get('uploaded_time', '')}"
        try:
            return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return datetime.min

    mode = sort_mode or "Newest first"
    if mode == "Oldest first":
        return sorted(results, key=lambda row: (datetime_key(row) == datetime.min, datetime_key(row)))
    if mode == "Cheapest first":
        return sorted(results, key=lambda row: (price_to_number(row.get("price", "")) is None, price_to_number(row.get("price", "")) or 0))
    if mode in {"Highest price first", "Most expensive first"}:
        return sorted(results, key=lambda row: price_to_number(row.get("price", "")) or -1, reverse=True)
    if mode == "Seller name A-Z":
        return sorted(results, key=lambda row: row.get("seller_name", ""))
    if mode == "City A-Z":
        return sorted(results, key=lambda row: row.get("city", ""))
    if mode == "Code A-Z":
        return sorted(results, key=lambda row: row.get("code", ""))
    return sorted(results, key=lambda row: datetime_key(row), reverse=True)


def parse_seller_profile_cards(soup: BeautifulSoup, seller_link: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for link in extract_listing_links_from_seller_page(soup):
        rows.append(row_from_listing_link(link, source_url=seller_link))
    return rows


def get_seller_plates(seller_link: str, max_pages: int = 3, timeout: int = 10) -> list[dict[str, str]]:
    if not seller_link:
        return []

    session = make_session()
    profile_cache: dict[str, tuple[str, str]] = {}
    listing_links: list[str] = []
    seen_links: set[str] = set()
    pages_to_fetch = [seller_link]
    seen_pages: set[str] = set()

    for page_url in pages_to_fetch[:max_pages]:
        if page_url in seen_pages:
            continue
        seen_pages.add(page_url)
        try:
            html = fetch_html(session, page_url, timeout=timeout)
        except requests.RequestException:
            continue

        soup = parse_soup(html)
        for extra_url in extract_pagination_urls(soup, page_url, max_pages=max_pages):
            if extra_url not in seen_pages and extra_url not in pages_to_fetch and len(pages_to_fetch) < max_pages:
                pages_to_fetch.append(extra_url)

        for row in parse_seller_profile_cards(soup, page_url):
            link = row.get("listing_link", "")
            if link and link not in seen_links:
                seen_links.add(link)
                listing_links.append(link)

    rows: list[dict[str, str]] = []
    for link in listing_links:
        try:
            detail = parse_listing_detail(link, session=session, profile_cache=profile_cache)
        except requests.RequestException:
            detail = row_from_listing_link(link, source_url=seller_link)
        rows.append(
            {
                "city": detail.get("city", ""),
                "plate_number": detail.get("plate_number", ""),
                "code": normalize_code(detail.get("code", "")),
                "price": detail.get("price", ""),
                "uploaded_date": detail.get("uploaded_date", "Not available"),
                "uploaded_time": detail.get("uploaded_time", "Not available"),
                "age_text": detail.get("age_text", "Not available"),
                "listing_link": detail.get("listing_link", link),
            }
        )

    return sort_results(rows, "Newest first")


def search_xplate(
    number: str,
    search_mode: str = "contains",
    max_price: str = "",
    min_price: str = "",
    cities: list[str] | None = None,
    number_format: str = "Any format",
    search_depth: str = "All pages",
    sort_mode: str = "Newest first",
    delay_seconds: float = 0,
    debug_callback: Callable[[str], None] | None = None,
    **_unused,
) -> list[dict[str, str]]:
    number = clean_text(number)
    mode = normalize_search_mode(search_mode)
    if not number and mode == "exact match":
        mode = "contains"
    selected_cities = cities or CITIES
    session = make_session()
    rows: list[dict[str, str]] = []
    global_seen_page_links: set[str] = set()
    max_pages = search_depth_to_max_pages(search_depth)

    debug_print(f"Selected format: {number_format}", debug_callback)
    debug_print(f"URL format value: {get_format_pattern(number_format) or '(none)'}", debug_callback)
    debug_print(f"Required digit length: {get_required_digit_length(number_format) or '(none)'}", debug_callback)

    for city in selected_cities:
        debug_print(f"Searching {city.title()}", debug_callback)
        city_page = 1
        while city_page <= max_pages:
            url = build_url(
                city,
                number,
                mode,
                max_price=max_price,
                min_price=min_price,
                selected_format=number_format,
                page=city_page,
            )
            debug_print(f"Scraping URL: {url}", debug_callback)
            try:
                html = fetch_html(session, url)
            except requests.RequestException as exc:
                debug_print(f"Search page failed for {city}, page {city_page}: {exc}", debug_callback)
                break

            soup = parse_soup(html)
            links = extract_listing_links_from_seller_page(soup)
            debug_print(f"Page {city_page} results: {len(links)}", debug_callback)
            if not links:
                break

            before_count = len(rows)
            new_links_on_page = 0
            for link in links:
                if link in global_seen_page_links:
                    continue
                global_seen_page_links.add(link)
                new_links_on_page += 1
                row = row_from_listing_link(link, source_url=url)
                if matches_search_mode(row["plate_number"], number, mode):
                    rows.append(row)
            debug_print(f"Loaded {len(rows)} total raw results so far", debug_callback)
            if new_links_on_page == 0:
                debug_print("No new listing links on this page; stopping pagination for this city", debug_callback)
                break

            if city_page >= max_pages:
                break
            if not has_next_page(soup, city_page) and len(rows) == before_count:
                break
            if not has_next_page(soup, city_page) and len(links) < 20:
                break
            city_page += 1

    unique_rows: list[dict[str, str]] = []
    seen_links: set[str] = set()
    for row in rows:
        link = row["listing_link"]
        if link in seen_links:
            continue
        seen_links.add(link)
        unique_rows.append(row)

    debug_print("Filtering results", debug_callback)
    if mode == "exact match":
        before_count = len(unique_rows)
        unique_rows = filter_exact_matches(unique_rows, number)
        debug_print(f"Exact-match filter removed {before_count - len(unique_rows)} rows", debug_callback)

    if number_format and number_format != "Any format":
        before_count = len(unique_rows)
        unique_rows = [row for row in unique_rows if matches_number_format(row.get("plate_number", ""), number_format)]
        debug_print(f"Final after local filtering: {len(unique_rows)}", debug_callback)
        debug_print(f"Number-format filter removed {before_count - len(unique_rows)} rows", debug_callback)

    debug_print("Collecting seller details", debug_callback)
    profile_cache: dict[str, tuple[str, str]] = {}
    final_rows: list[dict[str, str]] = []
    for row in unique_rows:
        try:
            details = parse_listing_detail(row["listing_link"], session=session, profile_cache=profile_cache)
            row.update({key: value for key, value in details.items() if value})
        except requests.RequestException as exc:
            debug_print(f"Detail page skipped: {row['listing_link']} ({exc})", debug_callback)

        final_rows.append(clean_result_row(row))
        if delay_seconds:
            time.sleep(delay_seconds)

    final_rows = apply_deal_rank(final_rows)
    final_rows = sort_results(final_rows, sort_mode)
    debug_print(f"Done. Final results: {len(final_rows)}", debug_callback)
    return final_rows


def save_results(results: list[dict[str, str]], number: str):
    output_dir = Path("results")
    output_dir.mkdir(exist_ok=True)

    df = pd.DataFrame(results)
    df = pd.DataFrame(columns=RESULT_COLUMNS) if df.empty else df.reindex(columns=RESULT_COLUMNS)

    safe_number = re.sub(r"[^0-9A-Za-z_-]+", "_", number).strip("_") or "search"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    csv_path = output_dir / f"xplate_results_{safe_number}_{timestamp}.csv"
    xlsx_path = output_dir / f"xplate_results_{safe_number}_{timestamp}.xlsx"

    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    df.to_excel(xlsx_path, index=False)
    format_excel_file(xlsx_path)
    return df, csv_path, xlsx_path


def format_excel_file(path: Path) -> None:
    workbook = load_workbook(path)
    sheet = workbook.active
    sheet.freeze_panes = "A2"
    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 55)
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Search Xplate plate numbers across UAE cities.")
    parser.add_argument("number", help="Plate number to search, for example 2007")
    parser.add_argument(
        "--mode",
        choices=["contains", "starts", "starts with", "ends", "ends with", "exact", "exact match"],
        default="contains",
    )
    parser.add_argument("--max-price", default="")
    parser.add_argument("--min-price", default="")
    parser.add_argument("--sort", default="Newest first")
    args = parser.parse_args()

    results = search_xplate(
        args.number,
        search_mode=args.mode,
        max_price=args.max_price,
        min_price=args.min_price,
        number_format="Any format",
        sort_mode=args.sort,
    )
    df, csv_path, xlsx_path = save_results(results, args.number)

    print("\nDone.")
    print(f"Total results: {len(df)}")
    print(f"CSV: {csv_path}")
    print(f"Excel: {xlsx_path}")
    if not df.empty:
        print(df.to_string(index=False))


if __name__ == "__main__":
    main()
